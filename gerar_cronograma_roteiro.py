# -*- coding: utf-8 -*-
"""
Gera cronograma dia a dia (2 vendedores) + roteiro de diagnóstico
Arquivo: kit_vendedor_higilabor.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = Workbook()

# ── cores ──────────────────────────────────────────────────────────
C_VERDE     = "1B5E20"
C_AZUL_ESC  = "0D2B5A"
COR = {
    "F1": ("1565C0", "BBDEFB"),   # azul — base atual
    "F2": ("B71C1C", "FFCDD2"),   # vermelho — NR-1
    "F3": ("1A237E", "C5CAE9"),   # azul escuro — telemedicina
    "F4": ("4A148C", "E1BEE7"),   # roxo — Pará
    "F5": ("1B5E20", "C8E6C9"),   # verde — FAP
    "FERIADO": ("757575", "EEEEEE"),
    "SEMANA":  ("0D2B5A", "E3F2FD"),
    "HDR":     ("0D2B5A", "FFFFFF"),
    "META":    ("E65100", "FFF3E0"),
}

def fill(cor): return PatternFill("solid", fgColor=cor)
def thin(c="CFD8DC"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def med_border(c="0D2B5A"):
    s = Side(style="medium", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, value, bg, fg="FFFFFF", bold=True, size=9,
        align="center", wrap=False, colspan=1):
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fg, name="Calibri")
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    c.border = thin("90A4AE")
    return c

def data_cell(ws, row, col, value, bg_esc, bg_clar, bold=False,
              size=9, wrap=True, colspan=1):
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=bg_esc, name="Calibri")
    c.fill = fill(bg_clar)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=wrap)
    c.border = thin()
    return c

# ══════════════════════════════════════════════════════════════════
# ABA 1 — CRONOGRAMA
# ══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Cronograma Diário"

# larguras
ws.column_dimensions["A"].width = 6   # #
ws.column_dimensions["B"].width = 10  # data
ws.column_dimensions["C"].width = 8   # dia
ws.column_dimensions["D"].width = 38  # V1
ws.column_dimensions["E"].width = 38  # V2
ws.column_dimensions["F"].width = 28  # Beto/Pará
ws.column_dimensions["G"].width = 22  # meta do dia

# ── título ────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
c = ws.cell(row=1, column=1, value="CRONOGRAMA DE CAMPANHA — HIGILABOR SST  |  Abr–Jun 2026")
c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
c.fill = fill("0D2B5A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:G2")
c = ws.cell(row=2, column=1,
            value="Vendedor 1: Frentes 1 (clientes) + 2 (NR-1 local)   |   "
                  "Vendedor 2: Frentes 3 (municípios) + 5 (FAP)   |   "
                  "Beto: Frente 4 (Mineração Pará)")
c.font = Font(bold=False, size=9, color="BBDEFB", name="Calibri")
c.fill = fill("1A237E")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

ws.row_dimensions[3].height = 6

# ── cabeçalho colunas ─────────────────────────────────────────
row = 4
ws.row_dimensions[row].height = 24
for col, (txt, bg) in enumerate(zip(
    ["Sem.", "Data", "Dia", "VENDEDOR 1 (F1+F2)", "VENDEDOR 2 (F3+F5)",
     "BETO — PARÁ (F4)", "META DO DIA"],
    ["0D2B5A"]*7), 1):
    hdr(ws, row, col, txt, bg, size=10, bold=True)
row += 1

# ── feriados ──────────────────────────────────────────────────
FERIADOS = {
    date(2026, 4, 20): "Tiradentes",
    date(2026, 5, 1): "Dia do Trabalhador",
}

# ── dados do cronograma ──────────────────────────────────────
# Cada entrada: (data, sem, v1_tarefas, v2_tarefas, beto_tarefas, meta)
DIAS = [
  # ── SEMANA 1 (08–10/abr) ──────────────────────────────────
  (date(2026,4,8),1,
   "[F1] WhatsApp para TODOS clientes ativos: 'Novidade — médica própria + NR-1 vence 26/mai'. Meta: 44 clientes abordados",
   "[F3] Preparar lista 32 municípios DDD 34 sem MT. WhatsApp para prefeituras de Centralina, Tupaciguara, Perdizes (já têm tel.)",
   "[F4] Disparar e-mail LOTE 1 (50 empresas A — ouro/cobre). Usar planilha acao_imediata_higilabor.xlsx",
   "44 clientes F1 | 3 municípios F3 | 50 e-mails F4"),

  (date(2026,4,9),1,
   "[F1] Follow-up ligação para quem não respondeu WhatsApp. [F2] Montar lista 30 empresas locais DDD 34 para NR-1",
   "[F3] WhatsApp prefeituras: Fronteira, Delta, Romaria, Carneirinho, Pirajuba, Planura",
   "[F4] Disparar e-mail LOTE 2 (45 empresas A restantes). Personalizar com nome da empresa e substância",
   "Ligações F1 | 6 municípios F3 | 45 e-mails F4"),

  (date(2026,4,10),1,
   "[F1] Enviar propostas para quem respondeu positivo. [F2] Iniciar WhatsApp lote 1 — 15 empresas locais NR-1",
   "[F3] WhatsApp direto para empresas nas cidades já abordadas (via indicação de prefeitura ou busca local)",
   "[F4] Montar lista visitas presenciais Beto — top 10 mineradoras (maior porte, melhor e-mail)",
   "Propostas F1 | 15 abordagens F2 | 10 visitas F4 planejadas"),

  # ── SEMANA 2 (13–17/abr) ──────────────────────────────────
  (date(2026,4,13),2,
   "[F2] WhatsApp lote 2 — 15 empresas locais NR-1. Setor: comércio e serviços Uberaba",
   "[F3] WhatsApp municípios: Ibiá, Nova Ponte, Santa Juliana, Campo Florido, Veríssimo",
   "[F4] Monitorar respostas e-mail lote 1 (enviado dia 8). Follow-up WhatsApp para quem abriu mas não respondeu",
   "15 F2 | 5 F3 | F4 follow-up"),

  (date(2026,4,14),2,
   "[F2] E-mail lote 1 NR-1 — 30 empresas locais. Assunto: '26 de maio: sua empresa adequou o PGR?'",
   "[F5] Iniciar abordagem FAP — 10 empresas. WhatsApp: 'Você sabe qual é o FAP de vocês?'",
   "[F4] E-mail follow-up lote 1 (e-mail 2 da sequência: urgência NR-1 para mineração)",
   "30 e-mails F2 | 10 WhatsApp F5 | F4 follow-up"),

  (date(2026,4,15),2,
   "[F2] Ligações para lista NR-1 — priorizar empresas sem resposta ao e-mail/WhatsApp",
   "[F3] WhatsApp empresas em: Matutina, São Francisco de Sales, União de Minas, Limeira do Oeste",
   "[F4] Beto: visitas presenciais às 3 primeiras mineradoras da lista (Marabá / Itaituba / Parauapebas)",
   "Ligações F2 | 4 F3 | 3 visitas F4"),

  (date(2026,4,16),2,
   "[F1] Follow-up para clientes que não fecharam na semana 1. Oferecer agendamento de diagnóstico gratuito",
   "[F5] E-mail lote FAP — 20 empresas. Assunto: 'Quanto sua empresa está pagando a mais de INSS?'",
   "[F4] Beto: visitas 4–6. Relato de campo para Octav — ajustar proposta conforme feedback",
   "F1 follow-up | 20 e-mails F5 | 3 visitas F4"),

  (date(2026,4,17),2,
   "[F2] WhatsApp lote 3 — 15 empresas locais NR-1. Setor: agro e indústria",
   "[F3] Consolidar respostas municípios semana 2. Enviar proposta F3 para interessados",
   "[F4] Monitorar respostas e-mail lote 2. Responder solicitações de proposta",
   "15 F2 | Propostas F3 | Propostas F4"),

  # ── FERIADO 20/abr ────────────────────────────────────────
  (date(2026,4,20),2, None, None, None, "FERIADO — Tiradentes"),

  # ── SEMANA 3 (21–24/abr) ──────────────────────────────────
  (date(2026,4,21),3,
   "[F2] Revisão pipeline: quantas propostas abertas? Ligar para todos. Fechar ou descartar",
   "[F3] WhatsApp municípios: Varjão de Minas, Guimarânia, Presidente Olegário, Rio Paranaíba",
   "[F4] Disparar e-mail ÚLTIMO CONTATO (e-mail 3) para quem não respondeu lotes 1+2",
   "Pipeline review | 4 F3 | F4 último contato"),

  (date(2026,4,22),3,
   "[F2] E-mail follow-up lote 1 NR-1 (e-mail 2: 'faltam 5 semanas'). Subir urgência",
   "[F5] Follow-up FAP — ligar para quem abriu e-mail mas não respondeu. Oferecer diagnóstico gratuito",
   "[F4] Beto: visitas 7–10. Foco em empresas que responderam e-mail com interesse",
   "F2 follow-up | Ligações F5 | 4 visitas F4"),

  (date(2026,4,23),3,
   "[F2] WhatsApp lote 4 — 20 novas empresas locais. Setor: saúde, educação, logística",
   "[F3] Abordagem empresas em municípios via secretaria de saúde + indicações de prefeitura",
   "[F4] Montar propostas para os leads confirmados de Beto. Enviar PDF em até 24h",
   "20 F2 | F3 indicações | Propostas F4"),

  (date(2026,4,24),3,
   "[F1] Check de propostas: quem está pendente há mais de 10 dias? Última tentativa",
   "[F5] Diagnósticos FAP gratuitos: realizar análise para quem confirmou interesse",
   "[F4] Acompanhar respostas + revisão pipeline Pará. Relatório para Octav",
   "Check F1 | Diagnósticos F5 | Relatório F4"),

  # ── SEMANA 4 (28/abr–02/mai) ──────────────────────────────
  (date(2026,4,28),4,
   "[F2] ⚠ FALTAM 4 SEMANAS — intensificar. E-mail urgência para toda lista. Subir tom",
   "[F3] WhatsApp: Araporã, Abadia dos Dourados, Cascalho Rico, Douradoquara, Grupiara",
   "[F4] Início de contratos fechados. Onboarding dos primeiros clientes Pará",
   "F2 urgência total | 5 F3 | Onboarding F4"),

  (date(2026,4,29),4,
   "[F2] Ligar PESSOALMENTE para top 10 leads quentes que ainda não fecharam",
   "[F5] Enviar plano de ação FAP para quem fez diagnóstico. Proposta formal",
   "[F4] Beto: follow-up pós-visitas — quem está considerando? Oferecer proposta revista",
   "Ligações F2 | Propostas F5 | F4 follow-up"),

  (date(2026,4,30),4,
   "[F2] Fechar todas as propostas possíveis antes de mai/26. Negociar condições",
   "[F3] Propostas formais para municípios com interesse confirmado",
   "[F4] Disparar nova rodada de e-mails para empresas B (média prioridade) do Pará",
   "Fechar F2 | Propostas F3 | F4 novo lote B"),

  # FERIADO 1/mai
  (date(2026,5,1),4, None, None, None, "FERIADO — Dia do Trabalhador"),

  # ── SEMANA 5 (05–09/mai) ──────────────────────────────────
  (date(2026,5,5),5,
   "[F2] ⚠⚠ FALTAM 21 DIAS — campanha em modo urgência. WhatsApp para TODO o pipeline aberto",
   "[F3] Abordagem final municípios sem resposta. Última chance antes do prazo",
   "[F4] Follow-up e-mail lote B Pará. Priorizar quem demonstrou interesse",
   "URGÊNCIA NR-1 | F3 final | F4 follow-up B"),

  (date(2026,5,6),5,
   "[F2] Oferecer entrega expressa: 'Garantimos entrega do PGR antes de 26/mai se assinar até 16/mai'",
   "[F5] Último push FAP: 'Fechar agora permite reduzir FAP ainda na apuração de 2026'",
   "[F4] Beto: nova rodada de visitas — leads quentes que não fecharam ainda",
   "Entrega expressa F2 | FAP fechamento | Visitas F4"),

  (date(2026,5,7),5,
   "[F2] Ligar para 100% dos leads abertos. Sem exceção. Meta: fechar ou descartar",
   "[F3] Consolidar contratos municípios. Iniciar PCMSO dos que fecharam",
   "[F4] Relatório intermediário: contratos fechados Pará vs. meta",
   "Ligações F2 | PCMSO F3 início | Relatório F4"),

  (date(2026,5,8),5,
   "[F2] Enviar propostas de última hora — validade 5 dias. Urgência máxima",
   "[F3] Onboarding clientes novos municípios. Coletar GHE e dados para PCMSO",
   "[F4] Novos e-mails para empresas B/C que não responderam anterior",
   "Urgência F2 | Onboarding F3 | F4 reativação"),

  (date(2026,5,9),5,
   "[F1+F2] Review semanal de resultados. Quantos fechados? Ajustar abordagem",
   "[F3+F5] Review semanal. Pipeline atual de municípios e FAP",
   "[F4] Review Pará com Beto. Quantos contratos? Ajustar plano semana 6",
   "REVIEW SEMANAL — Ajustar estratégia"),

  # ── SEMANA 6 (12–16/mai) ──────────────────────────────────
  (date(2026,5,12),6,
   "[F2] ⚠⚠⚠ FALTAM 14 DIAS. Última janela para entrega garantida. Ligar para todos",
   "[F3] Abordagem novos municípios com estratégia pós-NR-1: 'adeque antes que o fiscal chegue'",
   "[F4] Entrega dos primeiros laudos/PCMSO para clientes Pará assinados em abril",
   "14 DIAS NR-1 | F3 estratégia nova | Entrega F4"),

  (date(2026,5,13),6,
   "[F2] WhatsApp final — tom direto: 'Fechando agenda desta semana. Última vaga para garantir prazo'",
   "[F5] Diagnósticos FAP pendentes. Agendar apresentação dos resultados",
   "[F4] Beto: visitas de fechamento — foco em quem pediu proposta mas não assinou",
   "Agendas F2 | FAP diagnósticos | Fechamento F4"),

  (date(2026,5,14),6,
   "[F2] Assinaturas de contratos — prioridade absoluta hoje. Coleta de dados para elaboração PGR",
   "[F3] Assinaturas contratos municípios. Coleta de informações para PCMSO",
   "[F4] Assinaturas contratos Pará. Onboarding remoto",
   "ASSINATURAS — todos os contratos pendentes"),

  (date(2026,5,15),6,
   "[F2] Iniciar elaboração PGR para contratos assinados semana 6",
   "[F3] Elaboração PCMSO para novos clientes municípios",
   "[F4] Elaboração documentos Pará — PGR NR-22",
   "PRODUÇÃO — iniciar entregas"),

  (date(2026,5,16),6,
   "[F2] ⚠ PRAZO INTERNO: contratos assinados até hoje têm entrega garantida antes de 26/mai",
   "[F3+F5] Contratos de última hora — aceitar até hoje para garantir entrega",
   "[F4] Relatório quinzenal Beto — resultados e pipeline final",
   "ÚLTIMO DIA garantia de entrega antes 26/mai"),

  # ── SEMANA 7 (19–23/mai) ──────────────────────────────────
  (date(2026,5,19),7,
   "[F2] ENTREGA dos PGRs elaborados. Revisão com cliente. Assinatura do laudo",
   "[F3] ENTREGA PCMSOs — clientes municípios. Onboarding PCMSO telemedicina",
   "[F4] Continuar elaboração documentos Pará. Entregas parciais",
   "ENTREGAS semana 7"),

  (date(2026,5,20),7,
   "[F2] Entrega e revisão de laudos. Coletar feedback e testemunho do cliente",
   "[F3] Revisão e assinatura médica dos PCMSOs — dra. tia assina remotamente",
   "[F4] E-mail nova prospecção: empresas B Pará que ainda não abriram contato",
   "Entregas + testemunhos + nova prospecção"),

  (date(2026,5,21),7,
   "[F2] Pós-NR-1: iniciar abordagem 'adequação retroativa' para quem ficou de fora",
   "[F3] Abordagem novos municípios (expandir para DDD 33/35/38)",
   "[F4] Planejar expansão Pará fase 2 — novos municípios / estados mineradores",
   "Pós-NR-1: nova onda"),

  (date(2026,5,22),7,
   "[F2] Coletar depoimentos de clientes que adequaram. Preparar case para novos leads",
   "[F5] FAP: revisão de diagnósticos. Converter pendentes em contratos",
   "[F4] Relatório final campanha Pará para Octav",
   "Cases + FAP conversão + Relatório F4"),

  (date(2026,5,23),7,
   "[GERAL] Review semanal completo. Contagem final de contratos por frente",
   "[GERAL] Review + projeção MRR adicionado na campanha",
   "[F4] Beto: balanço final e plano de continuidade",
   "REVIEW FINAL PRÉ-PRAZO"),

  # ── SEMANA 8 (26–30/mai) ──────────────────────────────────
  (date(2026,5,26),8,
   "[NR-1 VIGENTE] Nova abordagem: 'A NR-1 já está em vigor. Sua empresa está irregular?'",
   "[F3] Expansão para cidades DDD 33 (177 municípios). Novo ciclo",
   "[F4] Iniciar prospecção fase 2 Pará — empresas B que não responderam",
   "NR-1 VIGENTE — nova onda de prospecção"),

  (date(2026,5,27),8,
   "[F2] Prospecção pós-prazo: empresas que ficaram sem adequar são leads quentes agora",
   "[F5] Novos diagnósticos FAP — empresas que conheceram a Higilabor pela campanha NR-1",
   "[F4] Novos e-mails Pará lote C",
   "Pós-prazo = nova oportunidade"),

  (date(2026,5,28),8,
   "[GERAL] Consolidar resultados finais da campanha 8 semanas",
   "[GERAL] DRE da campanha: receita gerada vs. meta",
   "[F4] Plano de ação Beto — mês 3",
   "CONSOLIDAÇÃO FINAL"),
]

# ── renderizar cronograma ─────────────────────────────────────
SEM_CORES = {
    1: ("0D2B5A","E3F2FD"), 2: ("1565C0","E8F5E9"),
    3: ("1B5E20","F3E5F5"), 4: ("6A1B9A","FCE4EC"),
    5: ("B71C1C","FFF8E1"), 6: ("E65100","E0F7FA"),
    7: ("00695C","FFF3E0"), 8: ("0D2B5A","F5F5F5"),
}
sem_atual = 0
lnum = 0

for idx, entrada in enumerate(DIAS):
    d, sem, v1, v2, beto, meta = entrada
    feriado = v1 is None

    # nova semana
    if sem != sem_atual:
        sem_atual = sem
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=7)
        c = ws.cell(row=row, column=1,
                    value=f"  SEMANA {sem}   —   "
                          f"{d.strftime('%d/%b').upper()}")
        esc, clar = SEM_CORES.get(sem, ("0D2B5A","E3F2FD"))
        c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        c.fill = fill(esc)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    lnum += 1
    ws.row_dimensions[row].height = 58 if not feriado else 18
    esc, clar = SEM_CORES.get(sem, ("0D2B5A","E3F2FD"))

    # col A — nº
    c = ws.cell(row=row, column=1, value=lnum)
    c.font = Font(bold=True, size=8, color="FFFFFF", name="Calibri")
    c.fill = fill(esc)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()

    # col B — data
    c = ws.cell(row=row, column=2, value=d.strftime("%d/%b"))
    c.font = Font(bold=True, size=9, color=esc, name="Calibri")
    c.fill = fill(clar)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()

    # col C — dia
    dias_pt = {0:"Seg",1:"Ter",2:"Qua",3:"Qui",4:"Sex",5:"Sáb",6:"Dom"}
    c = ws.cell(row=row, column=3, value=dias_pt[d.weekday()])
    c.font = Font(bold=False, size=9, color=esc, name="Calibri")
    c.fill = fill(clar)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()

    if feriado:
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        c = ws.cell(row=row, column=4, value=f"  {meta}")
        c.font = Font(bold=True, size=9, color="757575", name="Calibri")
        c.fill = fill("EEEEEE")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin()
    else:
        for col_idx, (txt, frente_cor) in enumerate(zip(
            [v1, v2, beto, meta],
            ["1565C0", "1B5E20", "4A148C", "E65100"]), 4):
            bg = "FFFFFF" if col_idx < 7 else "FFF3E0"
            c = ws.cell(row=row, column=col_idx, value=txt or "")
            c.font = Font(size=8, color=frente_cor, name="Calibri")
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True)
            c.border = thin()
            if col_idx == 7:
                c.font = Font(size=8, bold=True, color="E65100",
                              name="Calibri")
                c.fill = fill("FFF3E0")
    row += 1

# ── legenda ───────────────────────────────────────────────────
row += 1
ws.row_dimensions[row].height = 16
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
c = ws.cell(row=row, column=1,
            value="  LEGENDA:  [F1] Base atual (azul) · [F2] NR-1 local (vermelho) · "
                  "[F3] Municípios telemedicina (azul escuro) · [F4] Pará (roxo) · [F5] FAP (verde)")
c.font = Font(size=8, color="424242", name="Calibri")
c.fill = fill("ECEFF1")
c.alignment = Alignment(horizontal="left", vertical="center")

ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════
# ABA 2 — ROTEIRO DE DIAGNÓSTICO GRATUITO
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Roteiro de Diagnóstico")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 48
ws2.column_dimensions["D"].width = 40
ws2.column_dimensions["E"].width = 28

def rot_cell(ws, row, col, value, bg="FFFFFF", fg="212121",
             bold=False, size=9, wrap=True, align="left",
             colspan=1):
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fg, name="Calibri")
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="top",
                            wrap_text=wrap)
    c.border = thin()
    return c

def rot_sec(ws, row, texto, cor="0D2B5A"):
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=f"  {texto}")
    c.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    c.fill = fill(cor)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1

# título
ws2.merge_cells("A1:E1")
c = ws2.cell(row=1, column=1,
             value="ROTEIRO DE DIAGNÓSTICO GRATUITO — HIGILABOR SST")
c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
c.fill = fill("0D2B5A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

ws2.merge_cells("A2:E2")
c = ws2.cell(row=2, column=1,
             value="Guia para o vendedor conduzir a reunião/ligação de diagnóstico — "
                   "duração ideal: 20–30 minutos | Objetivo: identificar gaps e apresentar proposta")
c.font = Font(size=9, color="BBDEFB", name="Calibri")
c.fill = fill("1A237E")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[2].height = 16

ws2.row_dimensions[3].height = 6

# cabeçalho colunas
row = 4
ws2.row_dimensions[row].height = 20
for col, (txt, w) in enumerate(zip(
    ["#", "FASE / PERGUNTA", "O QUE INVESTIGAR / COMO PERGUNTAR",
     "SINAL DE ALERTA (GAP)", "COMO POSICIONAR A HIGILABOR"], [5,30,48,40,28]), 1):
    c = ws2.cell(row=row, column=col, value=txt)
    c.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
    c.fill = fill("0D2B5A")
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    c.border = thin()
row += 1

ROTEIRO = [
  # ── ABERTURA ──────────────────────────────────────────────
  ("ABERTURA (2–3 min)", None, None, None, None),
  ("1", "Apresentação pessoal",
   "'Bom dia [nome], sou [vendedor] da Higilabor SST de Uberaba. "
   "Trabalho com saúde e segurança do trabalho para empresas da região. "
   "Tenho uns 20 minutos com você hoje — vou fazer algumas perguntas "
   "rápidas para entender a situação de vocês e ver se consigo ajudar.'",
   "—",
   "Tom: consultor, não vendedor. Você está fazendo um diagnóstico."),
  ("2", "Contextualizar o motivo do contato",
   "[F2] 'A razão do meu contato é a NR-1 psicossocial — obrigatória em maio. "
   "Quero entender se vocês já foram orientados sobre isso.'\n"
   "[F3] 'Sei que em [cidade] não tem médico do trabalho local. "
   "A Higilabor resolveu isso por telemedicina.'\n"
   "[F5] 'Quero entender o FAP de vocês — pode ter economia significativa em INSS.'",
   "Se já sabem do prazo mas não agiram → LEAD QUENTE",
   "Se não sabem → você é a primeira informação que recebem → autoridade."),

  # ── DISCOVERY ─────────────────────────────────────────────
  ("DISCOVERY — SITUAÇÃO ATUAL (8–10 min)", None, None, None, None),
  ("3", "Quantos funcionários CLT?",
   "'Quantos funcionários vocês têm no total? Tem CLT, PJ ou terceirizados?'",
   "Se tem CLT → tem obrigação SST. Se tem terceirizados → PGR ainda se aplica.",
   "Usa para dimensionar o serviço e calcular o preço."),
  ("4", "Tem PGR ativo?",
   "'A empresa tem um Programa de Gerenciamento de Riscos (PGR) ativo hoje? "
   "Quando foi feito? Tem assinatura de engenheiro?'",
   "Sem PGR → autuação imediata em fiscalização → URGÊNCIA TOTAL\n"
   "PGR desatualizado (>1 ano) → precisa de revisão",
   "'Empresas sem PGR estão irregulares independente do porte. "
   "A Higilabor faz isso em 7 dias úteis.'"),
  ("5", "Tem PCMSO ativo?",
   "'E o PCMSO — Programa de Controle Médico? Quem é o médico coordenador? "
   "Quando foi feito o último exame periódico?'",
   "Sem PCMSO ou médico → irregularidade grave + base para ação trabalhista\n"
   "Médico parceiro externo que demora → oportunidade de substituição",
   "[F3] 'A Higilabor tem médica do trabalho própria — ASO por telemedicina, mais rápido e barato.'\n"
   "[F1] 'A Higilabor agora tem médica própria — não depende mais de terceiros.'"),
  ("6", "Situação NR-1 psicossocial",
   "'Vocês já foram orientados sobre a nova NR-1 psicossocial? "
   "O PGR de vocês já contempla avaliação de estresse e assédio?'",
   "Não sabe o que é → educar, criar urgência\n"
   "Sabe mas não agiu → LEAD QUENTE, fechar esta semana",
   "'A partir de 26 de maio, toda empresa precisa ter isso documentado. "
   "A gente faz em 7 dias úteis e garante entrega antes do prazo.'"),
  ("7", "[F5] Situação FAP",
   "'Você sabe qual é o FAP da empresa neste ano? "
   "Tem acesso ao portal do eSocial ou do INSS para verificar?'\n"
   "Se não sabe: 'Posso te ajudar a consultar agora — leva 2 minutos.'",
   "FAP > 1,0 → está pagando mais INSS → economia potencial imediata\n"
   "FAP = 2,0 com folha R$ 30k → pode economizar R$ 7.200/ano",
   "'Com o PCMSO e PGR certos, a próxima apuração do FAP já pode ser menor. "
   "Deixa eu calcular o impacto para vocês.'"),
  ("8", "Última fiscalização",
   "'A empresa já foi fiscalizada pelo MTE? Recebeu alguma notificação ou multa?'",
   "Sim → urgência total, não pode errar de novo\n"
   "Não → 'Depois da NR-1 entrar em vigor, as fiscalizações vão aumentar'",
   "Não alarmar — informar. 'Melhor regularizar antes do que durante fiscalização.'"),
  ("9", "Orçamento / decisão",
   "'Quem toma a decisão de contratar SST na empresa — é você ou tem mais alguém? "
   "Vocês têm orçamento reservado para isso?'",
   "Não é o decisor → pedir indicação para o decisor\n"
   "Sem orçamento → mostrar ROI (FAP, custo de multa, custo ação trabalhista)",
   "'O custo do serviço é muito menor do que o risco de uma autuação ou uma ação trabalhista. "
   "Posso te mostrar o cálculo?'"),

  # ── APRESENTAÇÃO ──────────────────────────────────────────
  ("APRESENTAÇÃO DA SOLUÇÃO (5–7 min)", None, None, None, None),
  ("10", "Resumo dos gaps identificados",
   "Antes de apresentar: 'Deixa eu resumir o que entendi: [repetir os gaps]. "
   "Estou certo?' → confirmar. Só depois apresentar.",
   "Se o cliente confirma os problemas → ele está comprando a necessidade",
   "Regra de ouro: apresente apenas a solução para os gaps que ele confirmou."),
  ("11", "Apresentar a Higilabor",
   "'A Higilabor resolve exatamente o que você descreveu:\n"
   "→ PGR e NR-1 psicossocial em 7 dias úteis\n"
   "→ PCMSO com médica própria — sem depender de parceiro\n"
   "[F3] → ASO por telemedicina — você não precisa deslocar ninguém\n"
   "[F5] → Plano de ação para reduzir FAP\n"
   "Temos engenheiro de segurança e médica do trabalho — entrega tudo na Higilabor.'",
   "—",
   "Máximo 3 minutos. Não detalhar tudo — deixar para a proposta escrita."),
  ("12", "Mostrar urgência concreta",
   "'Uma coisa importante: a NR-1 entra em vigor em 26 de maio — "
   "faltam [X] dias. Para garantir entrega antes do prazo, "
   "precisamos assinar ainda esta semana.'",
   "Se o cliente minimiza → 'Uma autuação do MTE começa em R$ 2.000 "
   "por infração por funcionário. Você tem [XX] funcionários.'",
   "Não ameaçar — informar o custo real do risco."),

  # ── FECHAMENTO ────────────────────────────────────────────
  ("FECHAMENTO (3–5 min)", None, None, None, None),
  ("13", "Pergunta de fechamento",
   "OPÇÃO A: 'Com base no que conversamos, faz sentido para você "
   "receber uma proposta formal esta semana?'\n\n"
   "OPÇÃO B: 'Você prefere que eu te mande a proposta por e-mail "
   "ou prefere que a gente veja juntos agora?'\n\n"
   "OPÇÃO C: 'Posso já te passar o valor para o porte de vocês — "
   "o que acha?'",
   "Se hesita → 'O que te impede de avançar agora?'\n"
   "Se é o preço → 'Posso comparar com o custo de não regularizar?'",
   "Sempre fechar com pergunta — nunca deixar a reunião sem um próximo passo definido."),
  ("14", "Definir próximo passo",
   "MÍNIMO: data para envio da proposta (em até 24h)\n"
   "IDEAL: data para assinatura do contrato\n"
   "ALTERNATIVA: agendar segunda reunião com o decisor\n\n"
   "'Você consegue me responder até [data] se avança ou não? "
   "Assim eu garanto o prazo de entrega.'",
   "Sem próximo passo definido = lead congelado",
   "Anotar no CRM imediatamente após a ligação."),
  ("15", "Objeções comuns e respostas",
   "OBJEÇÃO: 'Já tenho isso feito'\n"
   "→ 'Quando foi feito? Contempla a NR-1 psicossocial?'\n\n"
   "OBJEÇÃO: 'Está caro'\n"
   "→ 'Posso mostrar quanto uma autuação ou ação trabalhista custa? "
   "O serviço se paga em um único evento evitado.'\n\n"
   "OBJEÇÃO: 'Vou pensar'\n"
   "→ 'Claro. Mas o prazo é 26 de maio — posso te ligar amanhã para confirmar?'\n\n"
   "OBJEÇÃO: 'Tenho parceiro já'\n"
   "→ 'Ele já te apresentou a adequação para a NR-1 psicossocial? "
   "Posso fazer uma segunda opinião sem custo.'",
   "—",
   "Nunca discutir — sempre fazer uma pergunta que revele se o problema real é outro."),

  # ── PÓS-REUNIÃO ───────────────────────────────────────────
  ("PÓS-REUNIÃO — ações obrigatórias (até 2h após a reunião)", None, None, None, None),
  ("16", "Atualizar CRM",
   "Registrar: nome, empresa, CNPJ, WhatsApp, e-mail, "
   "situação atual (tem/não tem PGR, PCMSO), FAP, "
   "decisor (é o mesmo?), próximo passo, data follow-up",
   "CRM não atualizado = lead perdido",
   "Registrar imediatamente — memória falha."),
  ("17", "Enviar proposta em até 24h",
   "Usar template da frente correspondente (arquivo propostas_higilabor.xlsx). "
   "Preencher CNPJ, nome, porte, preços conforme tabela. "
   "Enviar por e-mail + PDF no WhatsApp.",
   "Proposta que demora mais de 24h perde urgência",
   "Se possível: apresentar o valor na própria reunião para acelerar decisão."),
  ("18", "Mensagem de follow-up WhatsApp",
   "'[Nome], conforme combinamos, segue a proposta em anexo. "
   "Prazo de entrega garantido se assinar até [data]. "
   "Qualquer dúvida me chama aqui.'",
   "—",
   "Simples, direto. Não escrever novela no follow-up."),
]

# ── renderizar roteiro ─────────────────────────────────────────
COR_FASES = {
    "ABERTURA": ("1565C0","E3F2FD"),
    "DISCOVERY": ("1B5E20","E8F5E9"),
    "APRESENTAÇÃO": ("4A148C","F3E5F5"),
    "FECHAMENTO": ("B71C1C","FCE4EC"),
    "PÓS-REUNIÃO": ("E65100","FFF3E0"),
}

num = 0
for entrada in ROTEIRO:
    idx, pergunta, como, alerta, posicionamento = entrada
    is_section = pergunta is None

    if is_section:
        ws2.row_dimensions[row].height = 22
        ws2.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=5)
        fase = idx.split("(")[0].strip().upper().split("—")[0].strip()
        esc, clar = next(
            ((v[0],v[1]) for k,v in COR_FASES.items() if k in fase),
            ("0D2B5A","FFFFFF"))
        c = ws2.cell(row=row, column=1, value=f"  ▶  {idx}")
        c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        c.fill = fill(esc)
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
    else:
        num += 1
        ws2.row_dimensions[row].height = 70
        # col A — número
        c = ws2.cell(row=row, column=1, value=num)
        c.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        c.fill = fill("37474F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        # col B — fase
        c = ws2.cell(row=row, column=2, value=pergunta)
        c.font = Font(bold=True, size=9, color="0D2B5A", name="Calibri")
        c.fill = fill("E8EAF6")
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
        c.border = thin()
        # col C — como perguntar
        c = ws2.cell(row=row, column=3, value=como)
        c.font = Font(size=8, color="212121", name="Calibri")
        c.fill = fill("FAFAFA")
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
        c.border = thin()
        # col D — sinal de alerta
        c = ws2.cell(row=row, column=4, value=alerta)
        c.font = Font(size=8, color="B71C1C", name="Calibri")
        c.fill = fill("FFF8E1")
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
        c.border = thin()
        # col E — posicionamento
        c = ws2.cell(row=row, column=5, value=posicionamento)
        c.font = Font(size=8, color=C_VERDE, name="Calibri")
        c.fill = fill("E8F5E9")
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
        c.border = thin()
        row += 1

# ── legenda ───────────────────────────────────────────────────
ws2.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════
# ABA 3 — METAS E KPIs
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Metas e KPIs")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 18
ws3.column_dimensions["F"].width = 22

ws3.merge_cells("A1:F1")
c = ws3.cell(row=1, column=1, value="METAS E KPIs — CAMPANHA HIGILABOR SST | 8 SEMANAS")
c.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
c.fill = fill("0D2B5A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

row = 3
# tabela geral
ws3.row_dimensions[row].height = 18
for col, txt in enumerate(["FRENTE","ABORDAGENS","PROPOSTAS","CONTRATOS","MRR ESTIMADO","PRAZO"], 1):
    c = ws3.cell(row=row, column=col, value=txt)
    c.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
    c.fill = fill("0D2B5A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
row += 1

metas = [
    ("F1 — Base Atual",        "44",  "10",  "5",  "R$ 1.500–4.000",  "Sem 1–2"),
    ("F2 — NR-1 Local",        "120", "30",  "10", "R$ 600–1.200",    "Sem 1–6"),
    ("F3 — Municípios",        "64",  "15",  "6",  "R$ 800–1.600",    "Sem 2–7"),
    ("F4 — Mineração Pará",    "95",  "14",  "4",  "R$ 2.000–5.000",  "Sem 1–8"),
    ("F5 — FAP",               "20",  "8",   "3",  "R$ 1.000–3.000",  "Sem 3–6"),
    ("TOTAL",                  "343", "77",  "28", "R$ 5.900–14.800", "8 sem."),
]

FRENTE_CORES = ["1565C0","B71C1C","1A237E","4A148C","1B5E20","0D2B5A"]
for i, (f, ab, pr, co, mrr, pz) in enumerate(metas):
    ws3.row_dimensions[row].height = 20
    destaque = (i == len(metas)-1)
    bg = "EEF2FF" if not destaque else "FFF9C4"
    cor = FRENTE_CORES[i]
    for col, val in enumerate([f, ab, pr, co, mrr, pz], 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = Font(bold=destaque, size=9 if not destaque else 10,
                      color=cor if col==1 else ("212121" if col!=4 else C_VERDE),
                      name="Calibri")
        c.fill = fill("FFF9C4" if destaque else ("FFFFFF" if i%2==0 else "F5F5F5"))
        c.alignment = Alignment(horizontal="center" if col>1 else "left",
                                vertical="center")
        c.border = thin()
    row += 1

row += 1
ws3.row_dimensions[row].height = 14
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
c = ws3.cell(row=row, column=1,
             value="  * MRR estimado por contrato/mês. Contratos anuais geram recorrência a partir do mês 13.")
c.font = Font(size=8, color="757575", name="Calibri")

row += 2
ws3.row_dimensions[row].height = 20
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
c = ws3.cell(row=row, column=1, value="  KPIs DIÁRIOS — VENDEDOR")
c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
c.fill = fill("1565C0")
c.alignment = Alignment(horizontal="left", vertical="center")
row += 1

for kpi in [
    ("Abordagens realizadas (WhatsApp + e-mail)", "≥ 15/dia", "contar ao final do dia"),
    ("Propostas enviadas", "≥ 2/dia", "registrar no CRM"),
    ("Contratos assinados", "meta: 1/dia na semana 5–6", "prioridade máxima"),
    ("Follow-ups no prazo (≤3 dias)", "100%", "nenhum lead sem follow-up"),
    ("CRM atualizado", "100% após cada contato", "registrar imediatamente"),
]:
    ws3.row_dimensions[row].height = 18
    for col, val in enumerate(kpi, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = Font(size=9, bold=(col==1), color=C_AZUL_ESC if col<3 else C_VERDE,
                      name="Calibri")
        c.fill = fill("E3F2FD" if row%2==0 else "FFFFFF")
        c.alignment = Alignment(horizontal="left" if col==1 else "center",
                                vertical="center")
        c.border = thin()
    ws3.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=2)
    row += 1

# ── salvar ────────────────────────────────────────────────────
OUT = r"C:\Users\octav\OneDrive\Apps\Claude\kit_vendedor_higilabor.xlsx"
wb.save(OUT)
print("Salvo:", OUT)
