# -*- coding: utf-8 -*-
"""
Gera propostas comerciais Higilabor — 5 frentes
Arquivo: propostas_higilabor.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

wb = Workbook()

# ── cores ────────────────────────────────────────────────────
C_AZUL_ESC  = "0D2B5A"
C_AZUL_MED  = "1565C0"
C_AZUL_CLAR = "BBDEFB"
C_VERDE     = "1B5E20"
C_VERDE_C   = "C8E6C9"
C_LARANJA   = "E65100"
C_LAR_C     = "FFE0B2"
C_CINZA     = "F5F5F5"
C_CINZA_ESC = "757575"
C_BRANCO    = "FFFFFF"
C_AMARELO   = "FFF9C4"
C_BORDA     = "CFD8DC"

def thin(cor=C_BORDA):
    s = Side(style='thin', color=cor)
    return Border(left=s, right=s, top=s, bottom=s)

def med(cor=C_AZUL_MED):
    s = Side(style='medium', color=cor)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(cor): return PatternFill("solid", fgColor=cor)

def fw(ws, col): return ws.column_dimensions[get_column_letter(col)].width

def set_col(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def cell(ws, row, col, value="", bold=False, size=10, color=C_AZUL_ESC,
         bg=None, align="left", valign="center", wrap=False,
         border=None, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color, italic=italic,
                  name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical=valign,
                            wrap_text=wrap)
    if bg:  c.fill = fill(bg)
    if border is not None: c.border = border
    return c

def merge(ws, r1, c1, r2, c2, value="", bold=False, size=11,
          color=C_BRANCO, bg=None, align="left", valign="center",
          wrap=False):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(bold=bold, size=size, color=color, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical=valign,
                            wrap_text=wrap)
    if bg: c.fill = fill(bg)
    return c

def cabecalho_proposta(ws, titulo, subtitulo, cor_titulo=C_AZUL_ESC,
                        cor_sub=C_AZUL_MED):
    """Cabeçalho padrão de proposta — 6 cols × 8 linhas"""
    ws.row_dimensions[1].height = 8
    # faixa azul escura
    for r in range(2, 7):
        for c_ in range(1, 7):
            ws.cell(row=r, column=c_).fill = fill(cor_titulo)
    merge(ws, 2, 1, 6, 5, "HIGILABOR SST", bold=True, size=20,
          color=C_BRANCO, bg=cor_titulo, align="left", valign="center")
    merge(ws, 2, 6, 6, 6, "Proposta\nComercial", bold=False, size=10,
          color="BBDEFB", bg=cor_titulo, align="right", valign="center")

    ws.row_dimensions[7].height = 5
    # faixa subtítulo
    for c_ in range(1, 7):
        ws.cell(row=8, column=c_).fill = fill(cor_sub)
    ws.row_dimensions[8].height = 22
    merge(ws, 8, 1, 8, 6, titulo, bold=True, size=13,
          color=C_BRANCO, bg=cor_sub, align="left", valign="center")

    ws.row_dimensions[9].height = 5
    ws.row_dimensions[10].height = 20
    for c_ in range(1, 7):
        ws.cell(row=10, column=c_).fill = fill("EBF5FB")
    merge(ws, 10, 1, 10, 6, subtitulo, bold=False, size=10,
          color=C_AZUL_ESC, bg="EBF5FB", align="left", valign="center",
          wrap=True)
    ws.row_dimensions[11].height = 8

def secao(ws, row, texto, cor=C_AZUL_MED):
    ws.row_dimensions[row].height = 20
    for c_ in range(1, 7):
        ws.cell(row=row, column=c_).fill = fill(cor)
    merge(ws, row, 1, row, 6, f"  {texto}", bold=True, size=11,
          color=C_BRANCO, bg=cor, align="left")
    return row + 1

def linha_dados(ws, row, label, valor, bg=C_CINZA, bg2=C_BRANCO):
    ws.row_dimensions[row].height = 18
    c1 = cell(ws, row, 1, label, bold=True, size=9, color=C_AZUL_ESC,
              bg=bg, border=thin())
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=2)
    c2 = cell(ws, row, 3, valor, bold=False, size=9, color="212121",
              bg=bg2, border=thin())
    ws.merge_cells(start_row=row, start_column=3,
                   end_row=row, end_column=6)
    ws.row_dimensions[row].height = 18
    return row + 1

def linha_item(ws, row, item, detalhe, status="Incluído",
               cor_status="1B5E20"):
    ws.row_dimensions[row].height = 30
    c1 = cell(ws, row, 1, item, bold=True, size=9, color=C_AZUL_ESC,
              bg=C_BRANCO, border=thin())
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=2)
    c2 = cell(ws, row, 3, detalhe, bold=False, size=9, color="424242",
              bg=C_BRANCO, border=thin(), wrap=True)
    ws.merge_cells(start_row=row, start_column=3,
                   end_row=row, end_column=5)
    c3 = cell(ws, row, 6, status, bold=True, size=9,
              color=cor_status, bg=C_VERDE_C, border=thin(), align="center")
    return row + 1

def linha_preco(ws, row, descricao, valor, destaque=False):
    bg = C_AMARELO if destaque else C_BRANCO
    ws.row_dimensions[row].height = 20
    c1 = cell(ws, row, 1, descricao, bold=destaque, size=9,
              color=C_AZUL_ESC, bg=bg, border=thin())
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=5)
    c2 = cell(ws, row, 6, valor, bold=destaque, size=10 if destaque else 9,
              color=C_VERDE if destaque else "424242", bg=bg,
              border=thin(), align="right")
    return row + 1

def espaco(ws, row, h=8):
    ws.row_dimensions[row].height = h
    return row + 1

def rodape(ws, row):
    row = espaco(ws, row, 10)
    ws.row_dimensions[row].height = 16
    for c_ in range(1, 7):
        ws.cell(row=row, column=c_).fill = fill(C_AZUL_ESC)
    merge(ws, row, 1, row, 3,
          "Higilabor SST  |  Uberaba/MG  |  (34) XXXX-XXXX",
          bold=False, size=8, color="90CAF9", bg=C_AZUL_ESC, align="left")
    merge(ws, row, 4, row, 6,
          "higilabor.com.br  |  higilabor@higilabor.com.br",
          bold=False, size=8, color="90CAF9", bg=C_AZUL_ESC, align="right")

def config_colunas(ws, widths):
    for i, w in enumerate(widths, 1):
        set_col(ws, i, w)

# ═══════════════════════════════════════════════════════════════
# PROPOSTA 1 — BASE ATUAL (clientes / ex-clientes)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "F1 – Revisão Completa"
config_colunas(ws1, [18, 18, 18, 18, 18, 18])

cabecalho_proposta(ws1,
    "PACOTE REVISÃO COMPLETA — CLIENTES HIGILABOR",
    "  Atualização do PGR com NR-1 Psicossocial + PCMSO com Médica do Trabalho Própria  |  Condições especiais para clientes ativos")

r = 12
r = secao(ws1, r, "DADOS DO CLIENTE")
r = linha_dados(ws1, r, "Empresa", "[Nome da Empresa]")
r = linha_dados(ws1, r, "CNPJ", "[XX.XXX.XXX/0001-XX]")
r = linha_dados(ws1, r, "Responsável", "[Nome do Contato]")
r = linha_dados(ws1, r, "Setor / Atividade", "[Ex: Agronegócio / Construção / Comércio]")
r = linha_dados(ws1, r, "Nº de Funcionários", "[XX] funcionários")
r = linha_dados(ws1, r, "Cidade", "[Cidade/MG]")
r = espaco(ws1, r)

r = secao(ws1, r, "CONTEXTO")
ws1.row_dimensions[r].height = 80
merge(ws1, r, 1, r, 6,
    "A Higilabor, como consultoria parceira, identificou duas necessidades imediatas para sua empresa:\n"
    "1. Adequação do PGR à NR-1 revisada — prazo regulatório: 26 de maio de 2026. Todo PGR deve contemplar avaliação de riscos psicossociais.\n"
    "2. Atualização do PCMSO — a Higilabor agora conta com Médica do Trabalho própria, eliminando dependência de parceiro externo e reduzindo o prazo de entrega.",
    bold=False, size=9, color="212121", bg="EBF5FB", align="left",
    valign="top", wrap=True)
r += 1
r = espaco(ws1, r)

r = secao(ws1, r, "ESCOPO DE SERVIÇOS")
ws1.row_dimensions[r].height = 16
merge(ws1, r, 1, r, 2, "Serviço", bold=True, size=9, color=C_BRANCO,
      bg=C_AZUL_MED, align="center")
merge(ws1, r, 3, r, 5, "Descrição", bold=True, size=9, color=C_BRANCO,
      bg=C_AZUL_MED, align="center")
cell(ws1, r, 6, "Status", bold=True, size=9, color=C_BRANCO,
     bg=C_AZUL_MED, align="center", border=thin())
r += 1
r = linha_item(ws1, r, "PGR Atualizado",
    "Revisão completa com inclusão de avaliação de riscos psicossociais (NR-1 revisada). Metodologia validada, laudo técnico assinado por Engenheiro de Segurança.")
r = linha_item(ws1, r, "Avaliação Psicossocial",
    "Diagnóstico de riscos de estresse, assédio, sobrecarga e violência — conforme exigência MTE. Inclui relatório gerencial e plano de ação.")
r = linha_item(ws1, r, "PCMSO",
    "Programa de Controle Médico de Saúde Ocupacional coordenado por Médica do Trabalho própria da Higilabor. Inclui planejamento de exames periódicos e admissionais.")
r = linha_item(ws1, r, "ASOs",
    "Emissão de Atestados de Saúde Ocupacional (admissional, periódico, demissional, mudança de função, retorno ao trabalho).")
r = linha_item(ws1, r, "Laudos Complementares",
    "Laudo de Insalubridade e/ou Periculosidade quando aplicável ao perfil de risco da empresa.", "Sob consulta", C_AZUL_MED)
r = espaco(ws1, r)

r = secao(ws1, r, "INVESTIMENTO")
r = linha_preco(ws1, r, "PGR Atualizado + NR-1 Psicossocial (entrega única)", "R$ [XXX]")
r = linha_preco(ws1, r, "PCMSO anual — coordenação médica (até [XX] func.)", "R$ [XXX]/ano")
r = linha_preco(ws1, r, "ASO unitário (por exame realizado)", "R$ [XX]/ASO")
r = linha_preco(ws1, r, "TOTAL PACOTE REVISÃO COMPLETA (12 meses)", "R$ [XXXX]", destaque=True)
r = espaco(ws1, r)

ws1.row_dimensions[r].height = 16
merge(ws1, r, 1, r, 6,
    "  ✓ Desconto de fidelidade aplicável para clientes com histórico Higilabor  |  Parcelamento em até 3x sem juros  |  Condições válidas até 10/05/2026",
    bold=False, size=8, color=C_VERDE, bg=C_VERDE_C, align="left")
r += 1
r = espaco(ws1, r)

r = secao(ws1, r, "PRAZO E CONDIÇÕES")
r = linha_dados(ws1, r, "Prazo de entrega (PGR + NR-1)", "Até 7 dias úteis após assinatura e envio das informações")
r = linha_dados(ws1, r, "Validade da proposta", "10 dias corridos")
r = linha_dados(ws1, r, "Pagamento", "50% na assinatura + 50% na entrega  |  ou à vista com desconto")
r = linha_dados(ws1, r, "Vigência do contrato PCMSO", "12 meses, renovável")
r = espaco(ws1, r)

r = secao(ws1, r, "ASSINATURA E ACEITE")
ws1.row_dimensions[r].height = 60
merge(ws1, r, 1, r, 3,
    "Higilabor SST\n\n_____________________________________\nOctav — Responsável Técnico",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center",
    valign="center")
merge(ws1, r, 4, r, 6,
    "[Empresa Contratante]\n\n_____________________________________\n[Nome] — [Cargo]",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center",
    valign="center")
r += 1
rodape(ws1, r)

# ═══════════════════════════════════════════════════════════════
# PROPOSTA 2 — NR-1 PSICOSSOCIAL (empresas locais)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("F2 – NR-1 Psicossocial")
config_colunas(ws2, [18, 18, 18, 18, 18, 18])

cabecalho_proposta(ws2,
    "ADEQUAÇÃO NR-1 PSICOSSOCIAL — PGR ATUALIZADO",
    "  Avaliação de Riscos Psicossociais + PGR em conformidade com a NR-1 revisada  |  Prazo regulatório: 26 de maio de 2026",
    cor_titulo="1B3A6B", cor_sub="1565C0")

r = 12
r = secao(ws2, r, "DADOS DO CLIENTE", cor="1565C0")
r = linha_dados(ws2, r, "Empresa", "[Nome da Empresa]")
r = linha_dados(ws2, r, "CNPJ", "[XX.XXX.XXX/0001-XX]")
r = linha_dados(ws2, r, "Responsável", "[Nome / Cargo]")
r = linha_dados(ws2, r, "Setor / CNAE", "[Ex: Comércio varejista / CNAE XXXX-X]")
r = linha_dados(ws2, r, "Nº de Funcionários", "[XX] funcionários (CLT)")
r = linha_dados(ws2, r, "Possui PGR atual?", "[Sim — data] / [Não]")
r = espaco(ws2, r)

r = secao(ws2, r, "POR QUE ESTA PROPOSTA É URGENTE", cor="B71C1C")
ws2.row_dimensions[r].height = 90
merge(ws2, r, 1, r, 6,
    "A NR-1 (Norma Regulamentadora nº 1) foi revisada pelo Ministério do Trabalho e Emprego.\n\n"
    "A partir de 26 de maio de 2026, o PGR de toda empresa com funcionários CLT deve contemplar:\n"
    "  → Identificação e avaliação de riscos psicossociais (estresse, assédio, sobrecarga, violência)\n"
    "  → Medidas de prevenção e controle documentadas\n"
    "  → Revisão periódica e registro no eSocial\n\n"
    "Empresas sem adequação ficam expostas a: autuação em fiscalização | uso como prova em ações trabalhistas | agravamento do FAP.",
    bold=False, size=9, color="212121", bg="FFEBEE", align="left",
    valign="top", wrap=True)
r += 1
r = espaco(ws2, r)

r = secao(ws2, r, "ESCOPO DE SERVIÇOS", cor="1565C0")
ws2.row_dimensions[r].height = 16
merge(ws2, r, 1, r, 2, "Serviço", bold=True, size=9, color=C_BRANCO,
      bg="1565C0", align="center")
merge(ws2, r, 3, r, 5, "Descrição", bold=True, size=9, color=C_BRANCO,
      bg="1565C0", align="center")
cell(ws2, r, 6, "Prazo", bold=True, size=9, color=C_BRANCO,
     bg="1565C0", align="center", border=thin())
r += 1

for item in [
    ("Diagnóstico Inicial",
     "Levantamento de riscos psicossociais existentes na empresa: função, jornada, ambiente, histórico de afastamentos.",
     "2 dias úteis"),
    ("Avaliação Psicossocial",
     "Aplicação de instrumento de avaliação validado. Relatório com identificação de fatores de risco e nível de exposição por setor.",
     "3 dias úteis"),
    ("Atualização do PGR",
     "Incorporação dos riscos psicossociais ao PGR existente. Revisão de medidas de prevenção e controle. Documento assinado por Engenheiro de Segurança.",
     "5 dias úteis"),
    ("Plano de Ação",
     "Elaboração de plano de ação com medidas corretivas e preventivas. Cronograma de implementação e indicadores de monitoramento.",
     "Incluso"),
    ("Treinamento (opcional)",
     "Capacitação de gestores e liderança sobre riscos psicossociais e obrigações legais. Presencial ou online.",
     "Sob consulta"),
]:
    ws2.row_dimensions[r].height = 30
    merge(ws2, r, 1, r, 2, item[0], bold=True, size=9, color=C_AZUL_ESC,
          bg=C_BRANCO, align="left")
    merge(ws2, r, 3, r, 5, item[1], bold=False, size=9, color="424242",
          bg=C_BRANCO, align="left", wrap=True)
    c = ws2.cell(row=r, column=6, value=item[2])
    c.font = Font(bold=True, size=9, color=C_VERDE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = fill(C_VERDE_C)
    c.border = thin()
    for col in range(1, 7):
        ws2.cell(row=r, column=col).border = thin()
    r += 1

r = espaco(ws2, r)
r = secao(ws2, r, "INVESTIMENTO", cor="1565C0")

ws2.row_dimensions[r].height = 18
merge(ws2, r, 1, r, 6,
    "  Tabela por porte de empresa — escolha o perfil mais próximo:",
    bold=True, size=9, color="1565C0", bg="EBF5FB")
r += 1

# tabela preços
ws2.row_dimensions[r].height = 18
for col, (txt, w) in enumerate(zip(
        ["Porte", "Funcionários", "PGR + NR-1 Psicossocial", "PCMSO incluído?", "Valor Total", "Prazo"],
        [20, 15, 30, 18, 18, 18]), 1):
    c = ws2.cell(row=r, column=col, value=txt)
    c.font = Font(bold=True, size=9, color=C_BRANCO, name="Calibri")
    c.fill = fill(C_AZUL_ESC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
r += 1

precos = [
    ("Micro",     "1–9",   "R$ 850",  "Não", "R$ 850",   "5 dias úteis"),
    ("Pequena",   "10–29", "R$ 1.200","Não", "R$ 1.200", "5 dias úteis"),
    ("Média",     "30–59", "R$ 1.800","Sim", "R$ 2.800", "7 dias úteis"),
    ("Grande",    "60–99", "R$ 2.400","Sim", "R$ 3.800", "7 dias úteis"),
]
bgs = [C_BRANCO, C_CINZA]
for i, p in enumerate(precos):
    ws2.row_dimensions[r].height = 18
    bg = bgs[i % 2]
    for col, val in enumerate(p, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.font = Font(size=9, bold=(col in [1,5]), color=C_AZUL_ESC
                      if col != 5 else C_VERDE, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill(bg)
        c.border = thin()
    r += 1

r = espaco(ws2, r)
ws2.row_dimensions[r].height = 16
merge(ws2, r, 1, r, 6,
    "  ✓ Proposta com condições especiais válida até 10/05/2026  |  Entrega garantida antes do prazo NR-1 de 26/05",
    bold=False, size=8, color=C_VERDE, bg=C_VERDE_C)
r += 1
r = espaco(ws2, r)

r = secao(ws2, r, "PRAZO E CONDIÇÕES", cor="1565C0")
r = linha_dados(ws2, r, "Prazo de entrega", "Até 7 dias úteis após recebimento das informações da empresa")
r = linha_dados(ws2, r, "Validade da proposta", "15 dias corridos")
r = linha_dados(ws2, r, "Pagamento", "50% na assinatura + 50% na entrega do laudo")
r = linha_dados(ws2, r, "Modalidade", "Presencial (Uberaba e região) ou por videoconferência + visita técnica")
r = espaco(ws2, r)

r = secao(ws2, r, "ASSINATURA E ACEITE", cor="1565C0")
ws2.row_dimensions[r].height = 60
merge(ws2, r, 1, r, 3,
    "Higilabor SST\n\n_____________________________________\nOctav — Engenheiro / Responsável",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center", valign="center")
merge(ws2, r, 4, r, 6,
    "[Empresa Contratante]\n\n_____________________________________\n[Nome] — [Cargo]",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center", valign="center")
r += 1
rodape(ws2, r)

# ═══════════════════════════════════════════════════════════════
# PROPOSTA 3 — PCMSO TELEMEDICINA (municípios sem MT)
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("F3 – PCMSO Telemedicina")
config_colunas(ws3, [18, 18, 18, 18, 18, 18])

cabecalho_proposta(ws3,
    "PCMSO POR TELEMEDICINA — EMPRESAS EM MUNICÍPIOS SEM MÉDICO DO TRABALHO LOCAL",
    "  Médica do Trabalho própria  |  ASO com validade legal plena (Res. CFM 2.314/2022)  |  Sem deslocamento de equipe",
    cor_titulo="1A237E", cor_sub="283593")

r = 12
r = secao(ws3, r, "DADOS DO CLIENTE", cor="283593")
r = linha_dados(ws3, r, "Empresa", "[Nome da Empresa]")
r = linha_dados(ws3, r, "Cidade", "[Município / MG]")
r = linha_dados(ws3, r, "CNPJ", "[XX.XXX.XXX/0001-XX]")
r = linha_dados(ws3, r, "Responsável", "[Nome / Cargo]")
r = linha_dados(ws3, r, "Nº de Funcionários", "[XX] funcionários")
r = linha_dados(ws3, r, "PCMSO atual", "[Tem / Não tem] — Último exame: [data]")
r = espaco(ws3, r)

r = secao(ws3, r, "SOLUÇÃO HIGILABOR — TELEMEDICINA", cor="283593")
ws3.row_dimensions[r].height = 80
merge(ws3, r, 1, r, 6,
    "Em municípios sem médico do trabalho registrado no CFM, a NR-7 permite que o PCMSO seja "
    "coordenado por médico de outra especialidade — ou por telemedicina quando não há profissional local.\n\n"
    "A Higilabor oferece:\n"
    "  → Médica do Trabalho própria que coordena e assina o PCMSO remotamente\n"
    "  → ASOs emitidos com validade legal plena (Res. CFM 2.314/2022 + RDC ANVISA 756/2022)\n"
    "  → Exames clínicos realizados com clínica parceira na região — a médica avalia e homologa via teleconsulta\n"
    "  → Redução de custo: sem diária de deslocamento de médico de fora",
    bold=False, size=9, color="212121", bg="E8EAF6", align="left",
    valign="top", wrap=True)
r += 1
r = espaco(ws3, r)

r = secao(ws3, r, "ESCOPO DE SERVIÇOS", cor="283593")
ws3.row_dimensions[r].height = 16
merge(ws3, r, 1, r, 2, "Serviço", bold=True, size=9, color=C_BRANCO,
      bg="283593", align="center")
merge(ws3, r, 3, r, 5, "Descrição", bold=True, size=9, color=C_BRANCO,
      bg="283593", align="center")
cell(ws3, r, 6, "Incluso", bold=True, size=9, color=C_BRANCO,
     bg="283593", align="center", border=thin())
r += 1

for item in [
    ("PCMSO — Elaboração",
     "Programa de Controle Médico de Saúde Ocupacional completo, elaborado e assinado pela Médica do Trabalho da Higilabor. Inclui cronograma de exames, fichas clínicas e relatório anual."),
    ("ASO — Admissional",
     "Atestado de Saúde Ocupacional para admissão. Exame clínico via teleconsulta + exames complementares conforme função."),
    ("ASO — Periódico",
     "Exames periódicos conforme periodicidade do PCMSO. Teleconsulta médica + complementares."),
    ("ASO — Demissional / Outros",
     "ASO demissional, retorno ao trabalho e mudança de função."),
    ("NR-1 Psicossocial",
     "Adequação do PGR com avaliação de riscos psicossociais. Integrado ao PCMSO.", ),
    ("Relatório Anual PCMSO",
     "Documento obrigatório com estatísticas de saúde, morbidade e análise epidemiológica."),
]:
    r = linha_item(ws3, r, item[0], item[1])

r = espaco(ws3, r)
r = secao(ws3, r, "INVESTIMENTO", cor="283593")

ws3.row_dimensions[r].height = 16
for col, txt in enumerate(["Porte","Func.","PCMSO Anual","ASO Unitário","Pacote Anual (PCMSO+NR-1)","Exames compl."], 1):
    c = ws3.cell(row=r, column=col, value=txt)
    c.font = Font(bold=True, size=9, color=C_BRANCO, name="Calibri")
    c.fill = fill(C_AZUL_ESC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
r += 1

for i, p in enumerate([
    ("Micro",   "1–9",   "R$ 600/ano",  "R$ 80",  "R$ 1.200",  "Sob consulta"),
    ("Pequena", "10–29", "R$ 1.000/ano","R$ 80",  "R$ 1.850",  "Sob consulta"),
    ("Média",   "30–59", "R$ 1.600/ano","R$ 80",  "R$ 2.800",  "Sob consulta"),
    ("Grande",  "60–99", "R$ 2.200/ano","R$ 80",  "R$ 3.800",  "Sob consulta"),
]):
    ws3.row_dimensions[r].height = 18
    for col, val in enumerate(p, 1):
        c = ws3.cell(row=r, column=col, value=val)
        c.font = Font(size=9, bold=(col in [1,5]), color=C_VERDE if col==5 else C_AZUL_ESC, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill(C_BRANCO if i%2==0 else C_CINZA)
        c.border = thin()
    r += 1

r = espaco(ws3, r)
ws3.row_dimensions[r].height = 16
merge(ws3, r, 1, r, 6,
    "  ✓ Exames complementares realizados com clínica parceira na região — sem deslocamento para Uberaba  "
    " |  Cobrança por ASO realizado — sem mensalidade fixa de exames",
    bold=False, size=8, color=C_VERDE, bg=C_VERDE_C)
r += 1
r = espaco(ws3, r)

r = secao(ws3, r, "PRAZO E CONDIÇÕES", cor="283593")
r = linha_dados(ws3, r, "Elaboração do PCMSO", "Até 7 dias úteis após assinatura e envio do GHE/funções")
r = linha_dados(ws3, r, "Validade da proposta", "20 dias corridos")
r = linha_dados(ws3, r, "Vigência do contrato", "12 meses, renovável automaticamente")
r = linha_dados(ws3, r, "Pagamento PCMSO", "Anual ou semestral")
r = linha_dados(ws3, r, "Pagamento ASOs", "Por evento realizado (faturamento mensal)")
r = espaco(ws3, r)

r = secao(ws3, r, "ASSINATURA E ACEITE", cor="283593")
ws3.row_dimensions[r].height = 60
merge(ws3, r, 1, r, 3,
    "Higilabor SST\n\n_____________________________________\nOctav / Dra. [Nome] — Médica do Trabalho",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center", valign="center")
merge(ws3, r, 4, r, 6,
    "[Empresa Contratante]\n\n_____________________________________\n[Nome] — [Cargo]",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center", valign="center")
r += 1
rodape(ws3, r)

# ═══════════════════════════════════════════════════════════════
# PROPOSTA 4 — MINERADORAS PARÁ
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("F4 – Mineração Pará")
config_colunas(ws4, [18, 18, 18, 18, 18, 18])

cabecalho_proposta(ws4,
    "PCMSO + PGR NR-22 + LTCAT — OPERAÇÕES DE MINERAÇÃO",
    "  Setor Mineral  |  Telemedicina para áreas remotas  |  NR-22 + NR-1 Psicossocial + NR-7",
    cor_titulo="4A148C", cor_sub="6A1B9A")

r = 12
r = secao(ws4, r, "DADOS DA EMPRESA MINERADORA", cor="6A1B9A")
r = linha_dados(ws4, r, "Empresa", "[Razão Social]")
r = linha_dados(ws4, r, "CNPJ", "[XX.XXX.XXX/0001-XX]")
r = linha_dados(ws4, r, "Município de operação", "[Município / PA]")
r = linha_dados(ws4, r, "Substância/Minério", "[Ex: Ouro / Cobre / Manganês]")
r = linha_dados(ws4, r, "Regime", "[Alvará de Pesquisa / Concessão de Lavra / Lavra Garimpeira]")
r = linha_dados(ws4, r, "Nº de trabalhadores", "[XX] trabalhadores diretos")
r = linha_dados(ws4, r, "Responsável SST atual", "[Nome / empresa] — ou [não possui]")
r = espaco(ws4, r)

r = secao(ws4, r, "OBRIGAÇÕES LEGAIS — MINERAÇÃO", cor="B71C1C")
ws4.row_dimensions[r].height = 80
merge(ws4, r, 1, r, 6,
    "Operações de mineração estão sujeitas a um conjunto mais exigente de normas SST:\n\n"
    "  NR-22 — Segurança e Saúde Ocupacional na Mineração: PGR específico para mineração, análise de risco de cada atividade, plano de emergência.\n"
    "  NR-7 — PCMSO: obrigatório para todos os trabalhadores, com monitoramento de exposição a silica, metais pesados, ruído e poeira.\n"
    "  LTCAT — Laudo Técnico das Condições Ambientais: exigido para fins previdenciários (aposentadoria especial).\n"
    "  NR-1 Psicossocial: obrigatória a partir de 26/05/2026 — inclui avaliação de fatores de estresse em áreas remotas e trabalho em turnos.",
    bold=False, size=9, color="212121", bg="FCE4EC", align="left",
    valign="top", wrap=True)
r += 1
r = espaco(ws4, r)

r = secao(ws4, r, "ESCOPO DE SERVIÇOS", cor="6A1B9A")
ws4.row_dimensions[r].height = 16
merge(ws4, r, 1, r, 2, "Serviço", bold=True, size=9, color=C_BRANCO,
      bg="6A1B9A", align="center")
merge(ws4, r, 3, r, 5, "Descrição", bold=True, size=9, color=C_BRANCO,
      bg="6A1B9A", align="center")
cell(ws4, r, 6, "Incluso", bold=True, size=9, color=C_BRANCO,
     bg="6A1B9A", align="center", border=thin())
r += 1

for item in [
    ("PGR — NR-22",
     "Programa de Gerenciamento de Riscos específico para mineração. Inventário de riscos por atividade, agentes físicos/químicos/biológicos, medidas de controle e monitoramento."),
    ("PCMSO — NR-7",
     "Elaborado e coordenado por Médica do Trabalho da Higilabor via telemedicina. Inclui monitoramento de silicose, saturnismo, perda auditiva e outros agravos específicos da mineração."),
    ("LTCAT",
     "Laudo Técnico das Condições Ambientais do Trabalho. Base para concessão de aposentadoria especial e enquadramento de insalubridade."),
    ("NR-1 Psicossocial",
     "Avaliação de riscos psicossociais integrada ao PGR. Inclui fatores de isolamento geográfico, trabalho em turnos e pressão de produção."),
    ("ASOs por Telemedicina",
     "Admissional, periódico e demissional. Teleconsulta médica com clínica parceira na região para exames complementares."),
    ("Treinamentos NR-22",
     "Treinamentos obrigatórios de Segurança em Mineração. Pode ser realizado in loco por parceiro ou online.", "Sob consulta", C_AZUL_MED),
]:
    args = item if len(item) == 4 else (*item, "Incluído", C_VERDE)
    r = linha_item(ws4, r, item[0], item[1],
                   item[2] if len(item) > 2 else "Incluído",
                   item[3] if len(item) > 3 else C_VERDE)

r = espaco(ws4, r)
r = secao(ws4, r, "INVESTIMENTO", cor="6A1B9A")
r = linha_preco(ws4, r, "PGR NR-22 (elaboração + implantação)", "R$ [X.XXX]")
r = linha_preco(ws4, r, "PCMSO anual (coordenação médica — por trabalhador)", "R$ [XXX]/trabalhador/ano")
r = linha_preco(ws4, r, "LTCAT", "R$ [XXX]")
r = linha_preco(ws4, r, "NR-1 Psicossocial (integrada ao PGR)", "R$ [XXX]")
r = linha_preco(ws4, r, "ASO unitário (por teleconsulta realizada)", "R$ [XX]/ASO")
r = linha_preco(ws4, r, "PACOTE ANUAL COMPLETO (PGR+PCMSO+LTCAT+NR-1)", "R$ [X.XXX]", destaque=True)
r = espaco(ws4, r)

ws4.row_dimensions[r].height = 16
merge(ws4, r, 1, r, 6,
    "  ✓ Atendimento 100% remoto — sem custo de deslocamento de equipe técnica para o Pará"
    "  |  Proposta válida por 20 dias  |  Início em até 5 dias úteis após assinatura",
    bold=False, size=8, color="6A1B9A", bg="F3E5F5")
r += 1
r = espaco(ws4, r)

r = secao(ws4, r, "PRAZO E CONDIÇÕES", cor="6A1B9A")
r = linha_dados(ws4, r, "PGR NR-22", "15–20 dias úteis (depende do porte e acesso ao local)")
r = linha_dados(ws4, r, "PCMSO + LTCAT", "10–15 dias úteis")
r = linha_dados(ws4, r, "Validade da proposta", "20 dias corridos")
r = linha_dados(ws4, r, "Vigência do contrato", "12 meses")
r = linha_dados(ws4, r, "Pagamento", "30% assinatura + 40% entrega PGR + 30% entrega PCMSO/LTCAT")
r = espaco(ws4, r)

r = secao(ws4, r, "ASSINATURA E ACEITE", cor="6A1B9A")
ws4.row_dimensions[r].height = 60
merge(ws4, r, 1, r, 3,
    "Higilabor SST\n\n_____________________________________\nOctav / Dra. [Nome] — Médica do Trabalho",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center", valign="center")
merge(ws4, r, 4, r, 6,
    "[Empresa Mineradora]\n\n_____________________________________\n[Nome] — [Cargo]",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center", valign="center")
r += 1
rodape(ws4, r)

# ═══════════════════════════════════════════════════════════════
# PROPOSTA 5 — FAP (redução custo previdenciário)
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("F5 – FAP")
config_colunas(ws5, [18, 18, 18, 18, 18, 18])

cabecalho_proposta(ws5,
    "GESTÃO DO FAP — REDUÇÃO DO CUSTO PREVIDENCIÁRIO",
    "  Diagnóstico + Plano de Ação para redução do Fator Acidentário de Prevenção (FAP)  |  Economia mensurável em INSS",
    cor_titulo=C_VERDE, cor_sub="2E7D32")

r = 12
r = secao(ws5, r, "DADOS DO CLIENTE", cor="2E7D32")
r = linha_dados(ws5, r, "Empresa", "[Nome da Empresa]")
r = linha_dados(ws5, r, "CNPJ", "[XX.XXX.XXX/0001-XX]")
r = linha_dados(ws5, r, "Responsável / DP", "[Nome / Cargo]")
r = linha_dados(ws5, r, "Nº de Funcionários", "[XX] funcionários")
r = linha_dados(ws5, r, "Folha de pagamento (estimada)", "R$ [XX.XXX]/mês")
r = linha_dados(ws5, r, "FAP atual (se conhecido)", "[X,XX] — ou [não sabe]")
r = linha_dados(ws5, r, "Alíquota RAT", "[1% / 2% / 3%]")
r = espaco(ws5, r)

r = secao(ws5, r, "O IMPACTO FINANCEIRO DO FAP", cor="E65100")
ws5.row_dimensions[r].height = 20
merge(ws5, r, 1, r, 6,
    "  Simulação de impacto — Empresa com folha de R$ 30.000/mês e RAT de 2%:",
    bold=True, size=9, color="E65100", bg="FFF3E0")
r += 1

ws5.row_dimensions[r].height = 16
for col, txt in enumerate(["FAP","RAT Ajustado","Custo INSS/mês","Custo INSS/ano","Diferença anual vs. FAP 1,0"], 1):
    c = ws5.cell(row=r, column=col, value=txt)
    c.font = Font(bold=True, size=9, color=C_BRANCO, name="Calibri")
    c.fill = fill(C_AZUL_ESC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
ws5.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
r += 1

for i, (fap, rat, mes, ano, dif, destaque) in enumerate([
    ("0,50", "1,0%", "R$ 300",  "R$ 3.600",  "– R$ 7.200", False),
    ("1,00", "2,0%", "R$ 600",  "R$ 7.200",  "referência",  True),
    ("1,50", "3,0%", "R$ 900",  "R$ 10.800", "+ R$ 3.600", False),
    ("2,00", "4,0%", "R$ 1.200","R$ 14.400", "+ R$ 7.200", False),
]):
    ws5.row_dimensions[r].height = 18
    for col, val in enumerate([fap, rat, mes, ano, dif, ""], 1):
        c = ws5.cell(row=r, column=col, value=val)
        c.font = Font(size=9, bold=destaque,
                      color=C_VERDE if not destaque else C_AZUL_ESC,
                      name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill(C_AMARELO if destaque else (C_BRANCO if i%2==0 else C_CINZA))
        c.border = thin()
    if dif not in ["referência", ""]:
        ws5.cell(row=r, column=5).font = Font(
            size=9, bold=True,
            color=C_VERDE if "–" in dif else "C62828", name="Calibri")
    ws5.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    r += 1
r = espaco(ws5, r)

r = secao(ws5, r, "ESCOPO DE SERVIÇOS", cor="2E7D32")
ws5.row_dimensions[r].height = 16
merge(ws5, r, 1, r, 2, "Fase", bold=True, size=9, color=C_BRANCO, bg="2E7D32", align="center")
merge(ws5, r, 3, r, 5, "Descrição", bold=True, size=9, color=C_BRANCO, bg="2E7D32", align="center")
cell(ws5, r, 6, "Prazo", bold=True, size=9, color=C_BRANCO, bg="2E7D32", align="center", border=thin())
r += 1

for item in [
    ("1 — Diagnóstico FAP",
     "Análise do NTEP, histórico de CATs, afastamentos e benefícios previdenciários. Cálculo da contribuição atual e projeção de redução possível. GRATUITO para empresas com 10+ funcionários.",
     "3 dias úteis"),
    ("2 — Plano de Ação SST",
     "Estruturação ou revisão do PCMSO e PGR com foco em redução de afastamentos e CATs. Inclui protocolo de notificação correta de acidentes.",
     "10 dias úteis"),
    ("3 — Treinamento Gestores",
     "Capacitação do RH/DP sobre registro correto de CATs, gestão de afastamentos e interface com o eSocial para minimizar impacto no FAP.",
     "2h — agendável"),
    ("4 — Monitoramento Anual",
     "Acompanhamento semestral do indicador FAP + relatório de evolução + ajustes no PCMSO conforme resultados.",
     "Anual"),
]:
    ws5.row_dimensions[r].height = 30
    merge(ws5, r, 1, r, 2, item[0], bold=True, size=9, color=C_AZUL_ESC,
          bg=C_BRANCO, align="left")
    merge(ws5, r, 3, r, 5, item[1], bold=False, size=9, color="424242",
          bg=C_BRANCO, align="left", wrap=True)
    c = ws5.cell(row=r, column=6, value=item[2])
    c.font = Font(bold=True, size=9, color=C_VERDE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = fill(C_VERDE_C)
    c.border = thin()
    for col in range(1, 7):
        ws5.cell(row=r, column=col).border = thin()
    r += 1

r = espaco(ws5, r)
r = secao(ws5, r, "INVESTIMENTO", cor="2E7D32")
r = linha_preco(ws5, r, "Fase 1 — Diagnóstico FAP (gratuito para 10+ func.)", "GRATUITO")
r = linha_preco(ws5, r, "Fase 2 — Plano de ação SST (PCMSO + PGR com foco FAP)", "R$ [X.XXX]")
r = linha_preco(ws5, r, "Fase 3 — Treinamento gestores (por turma)", "R$ [XXX]")
r = linha_preco(ws5, r, "Fase 4 — Monitoramento anual", "R$ [XXX]/ano")
r = linha_preco(ws5, r, "PACOTE COMPLETO FAP (Fases 2+3+4)", "R$ [X.XXX]", destaque=True)
r = espaco(ws5, r)

ws5.row_dimensions[r].height = 28
merge(ws5, r, 1, r, 6,
    "  ROI estimado: para empresa com folha R$ 30k/mês e FAP 2,0 → redução para 1,0 representa economia de R$ 7.200/ano em INSS.\n"
    "  O pacote completo se paga em menos de [X] meses com a redução gerada.",
    bold=False, size=9, color=C_VERDE, bg=C_VERDE_C, wrap=True)
r += 1
r = espaco(ws5, r)

r = secao(ws5, r, "PRAZO E CONDIÇÕES", cor="2E7D32")
r = linha_dados(ws5, r, "Diagnóstico gratuito", "Até 3 dias úteis após envio dos dados do eSocial/CNIS")
r = linha_dados(ws5, r, "Validade da proposta", "15 dias corridos")
r = linha_dados(ws5, r, "Pagamento", "50% na assinatura do Plano de Ação + 50% na entrega")
r = espaco(ws5, r)

r = secao(ws5, r, "ASSINATURA E ACEITE", cor="2E7D32")
ws5.row_dimensions[r].height = 60
merge(ws5, r, 1, r, 3,
    "Higilabor SST\n\n_____________________________________\nOctav — Responsável Técnico",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center", valign="center")
merge(ws5, r, 4, r, 6,
    "[Empresa Contratante]\n\n_____________________________________\n[Nome] — [Cargo]",
    bold=False, size=9, color="424242", bg=C_BRANCO, align="center", valign="center")
r += 1
rodape(ws5, r)

# ── salvar ───────────────────────────────────────────────────
OUT = r"C:\Users\octav\OneDrive\Apps\Claude\propostas_higilabor.xlsx"
wb.save(OUT)
print("Salvo:", OUT)
