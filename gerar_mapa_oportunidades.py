"""
Gera planilha de mapeamento de oportunidades PCMSO telemedicina
Combina 853 MUNICÍPIOS.xlsx com classificação competitiva de médicos do trabalho
"""
import os, sys, unicodedata

def norm(s):
    """Normaliza: uppercase sem acentos para comparação"""
    if not s:
        return ""
    s = str(s).strip().upper()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# 1. CLASSIFICAÇÃO COMPETITIVA (pesquisa realizada)
# ─────────────────────────────────────────────────────────────
# Formato: "MUNICÍPIO": ("STATUS", "CONCORRENTE / OBSERVAÇÃO")
CLASSIFICACAO = {
    # ── SATURADO ─ evitar ──────────────────────────────────────
    "Uberaba":              ("SATURADO",              "25+ médicos MT, múltiplas clínicas (SESI, Mérito, Asseme, WTA…)"),
    "Uberlândia":           ("SATURADO",              "8+ clínicas (Mednet, Ubermed, Sermed, M7, Asseme, WTA, Pró Saúde, WR)"),
    "Patos de Minas":       ("SATURADO",              "GESAT municipal + clínicas privadas (Alfa Consultoria, outros)"),
    "Patrocínio":           ("SATURADO",              "Protege Minas — SEDE em Patrocínio"),
    "Monte Carmelo":        ("SATURADO",              "Protege Minas — filial Monte Carmelo"),
    "Araxá":                ("SATURADO",              "Clínica local confirmada (pesquisa anterior)"),
    "Ituiutaba":            ("SATURADO",              "Lider Medicina do Trabalho — sede"),
    "Araguari":             ("SATURADO",              "ACACIA Medicina Ocupacional"),

    # ── COM CONCORRENTE LOCAL ─ entrada difícil ────────────────
    "Sacramento":           ("COM CONCORRENTE",       "Sol-Saúde Ocupacional (local, TeleListas)"),
    "Serra do Salitre":     ("COM CONCORRENTE",       "Protege Minas — filial Serra do Salitre"),
    "Conceição das Alagoas":("COM CONCORRENTE",       "Dr. Luciano (médico local)"),
    "Capinópolis":          ("COM CONCORRENTE",       "Interclínica"),
    "Coromandel":           ("COM CONCORRENTE",       "Mérito + Expertisa + MGM (regionais itinerantes)"),
    "São Gotardo":          ("COM CONCORRENTE",       "Protege Minas + Mérito (regionais)"),
    "Carmo do Paranaíba":   ("COM CONCORRENTE",       "Mérito Medicina Ocupacional (regional)"),

    # ── CONCORRENTE ITINERANTE ─ entrada possível ──────────────
    "Campina Verde":        ("ITINERANTE",            "Lider Medicina (Ituiutaba) — atende itinerante"),
    "Canápolis":            ("ITINERANTE",            "Lider Medicina (Ituiutaba) — atende itinerante"),
    "Frutal":               ("ITINERANTE",            "Lider Medicina (Ituiutaba) — atende itinerante"),
    "Gurinhatã":            ("ITINERANTE",            "Lider Medicina (Ituiutaba) — atende itinerante"),
    "Itapagipe":            ("ITINERANTE",            "Lider Medicina (Ituiutaba) — atende itinerante"),
    "Monte Alegre de Minas":("ITINERANTE",            "Lider Medicina (Ituiutaba) — atende itinerante"),
    "Prata":                ("ITINERANTE",            "Lider Medicina (Ituiutaba) — atende itinerante"),
    "Santa Vitória":        ("ITINERANTE",            "Lider Medicina (Ituiutaba) — atende itinerante"),
    "Iturama":              ("ITINERANTE",            "Lider Medicina (Pontal do Triângulo — região)"),
    "Lagoa Formosa":        ("ITINERANTE",            "Coberta por provedores de Patos de Minas"),
    "João Pinheiro":        ("ITINERANTE",            "Provavelmente coberto por itinerante regional — verificar"),
    "Lagamar":              ("ITINERANTE",            "Área de influência Patos de Minas — verificar"),

    # ── ALTA OPORTUNIDADE ─ sem provedor conhecido ────────────
    "Centralina":           ("OPORTUNIDADE ALTA",     "Já cliente Higilabor ✓ — sem MT local"),
    "Araporã":              ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Cascalho Rico":        ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Delta":                ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Fronteira":            ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Abadia dos Dourados":  ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "Douradoquara":         ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade muito pequena"),
    "Grupiara":             ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade muito pequena"),
    "Guimarânia":           ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "Varjão de Minas":      ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "Matutina":             ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Ibiá":                 ("OPORTUNIDADE ALTA",     "Sem provedor identificado — verificar"),
    "Pedrinópolis":         ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Perdizes":             ("OPORTUNIDADE ALTA",     "Sem provedor identificado — verificar"),
    "Nova Ponte":           ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "Romaria":              ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Iraí de Minas":        ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "Santa Juliana":        ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "Tapira":               ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Campo Florido":        ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "Pirajuba":             ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Planura":              ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "São Francisco de Sales":("OPORTUNIDADE ALTA",   "Sem provedor identificado"),
    "Comendador Gomes":     ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Veríssimo":            ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Limeira do Oeste":     ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "União de Minas":       ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Carneirinho":          ("OPORTUNIDADE ALTA",     "Sem provedor identificado"),
    "Ipiaçu":               ("OPORTUNIDADE ALTA",     "Sem provedor identificado — cidade pequena"),
    "Tupaciguara":          ("OPORTUNIDADE ALTA",     "Sem provedor identificado — verificar CFM"),
    "Presidente Olegário":  ("OPORTUNIDADE ALTA",     "Sem provedor identificado — verificar"),
    "Rio Paranaíba":        ("OPORTUNIDADE ALTA",     "Sem provedor identificado — verificar"),
}

# Ordem de prioridade para ordenação
ORDEM_STATUS = {
    "OPORTUNIDADE ALTA": 1,
    "ITINERANTE": 2,
    "COM CONCORRENTE": 3,
    "SATURADO": 4,
    "NÃO MAPEADO": 5,
}

# Cores por status
CORES = {
    "OPORTUNIDADE ALTA": "00C851",   # verde
    "ITINERANTE":         "FFBB33",   # amarelo
    "COM CONCORRENTE":    "FF6D00",   # laranja
    "SATURADO":           "CC0000",   # vermelho
    "NÃO MAPEADO":        "AAAAAA",   # cinza
}

# ─────────────────────────────────────────────────────────────
# 2. LER 853 MUNICÍPIOS.xlsx
# ─────────────────────────────────────────────────────────────
ONEDRIVE = r"C:\Users\octav\OneDrive\DOCS OCTAVIO"
municipios_file = None
for f in os.listdir(ONEDRIVE):
    if "MUNIC" in f.upper() and f.endswith(".xlsx"):
        municipios_file = os.path.join(ONEDRIVE, f)
        break

if not municipios_file:
    print("ERRO: arquivo 853 MUNICÍPIOS não encontrado em", ONEDRIVE)
    sys.exit(1)

print("Lendo arquivo de municipios...")
wb_in = openpyxl.load_workbook(municipios_file, read_only=True, data_only=True)
ws_in = wb_in.active

rows_raw = []
for i, row in enumerate(ws_in.iter_rows(values_only=True)):
    if i == 0:
        continue  # cabeçalho
    r = list(row) + [None] * (11 - len(list(row)))
    rows_raw.append(r)

wb_in.close()
print(f"  {len(rows_raw)} municípios lidos")

# ─────────────────────────────────────────────────────────────
# 3. ENRIQUECER COM CLASSIFICAÇÃO E FILTRAR DDD 34
# ─────────────────────────────────────────────────────────────
# Colunas: #, Município, Prefeito, Cel.Prefeito, Tel.Município, Email, Partido, _, _, DATAS, SITUAÇÃO

DDD34_PREFIXES = [
    "34", "(34)", "034",
]

def is_ddd34(tel):
    if not tel:
        return False
    s = str(tel).strip().replace(" ", "").replace("-", "")
    return s.startswith("34") or s.startswith("(34)") or s.startswith("034")

registros = []
for r in rows_raw:
    num       = r[0]
    municipio = str(r[1]).strip() if r[1] else ""
    prefeito  = str(r[2]).strip() if r[2] else ""
    cel_pref  = str(r[3]).strip() if r[3] else ""
    tel_mun   = str(r[4]).strip() if r[4] else ""
    email     = str(r[5]).strip() if r[5] else ""
    partido   = str(r[6]).strip() if r[6] else ""
    situacao  = str(r[10]).strip() if r[10] else ""

    if not municipio or municipio == "None":
        continue

    # detectar DDD
    ddd = ""
    for tel in [cel_pref, tel_mun]:
        s = str(tel).replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
        if len(s) >= 2:
            ddd = s[:2]
            break

    # lookup normalizado
    status, observacao = ("NÃO MAPEADO", "Verificar CFM manualmente")
    mun_norm = norm(municipio)
    for k, v in CLASSIFICACAO.items():
        if norm(k) == mun_norm:
            status, observacao = v
            break

    registros.append({
        "num": num,
        "municipio": municipio,
        "prefeito": prefeito,
        "cel_prefeito": cel_pref,
        "tel_municipio": tel_mun,
        "email": email,
        "partido": partido,
        "ddd": ddd,
        "status": status,
        "observacao": observacao,
        "situacao": situacao,
    })

# Separar DDD 34 e restante
ddd34 = [r for r in registros if r["ddd"] == "34"]
outros = [r for r in registros if r["ddd"] != "34"]

# Ordenar por prioridade
def sort_key(r):
    return (ORDEM_STATUS.get(r["status"], 5), r["municipio"])

ddd34.sort(key=sort_key)
todos = sorted(registros, key=sort_key)

print(f"  DDD 34: {len(ddd34)} municípios")
print(f"  Outros DDDs: {len(outros)} municípios")

# ─────────────────────────────────────────────────────────────
# 4. GERAR EXCEL DE SAÍDA
# ─────────────────────────────────────────────────────────────
OUTPUT = r"C:\Users\octav\OneDrive\Apps\Claude\mapa_oportunidades_pcmso.xlsx"

wb = Workbook()

def thin_border():
    thin = Side(style='thin', color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def escrever_aba(ws, dados, titulo):
    ws.title = titulo

    # ── Cabeçalho ────────────────────────────────────────────
    headers = [
        "#", "Município", "Status MT", "Prefeito",
        "Cel. Prefeito", "Tel. Município", "Email Pref.", "DDD",
        "Observação Competitiva", "Partido"
    ]
    col_widths = [5, 28, 22, 24, 18, 18, 32, 6, 48, 12]

    hdr_fill = PatternFill("solid", fgColor="1A237E")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30

    # ── Dados ─────────────────────────────────────────────────
    opp_count = 0
    for i, r in enumerate(dados, 2):
        status = r["status"]
        cor = CORES.get(status, "AAAAAA")
        row_fill = PatternFill("solid", fgColor=cor + "22")  # leve tint
        status_fill = PatternFill("solid", fgColor=cor)
        status_font = Font(bold=True, color="FFFFFF", size=9)

        vals = [
            r["num"], r["municipio"], r["status"], r["prefeito"],
            r["cel_prefeito"], r["tel_municipio"], r["email"], r["ddd"],
            r["observacao"], r["partido"]
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 9))
            if col == 3:  # coluna status
                cell.fill = status_fill
                cell.font = status_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = row_fill
                cell.font = Font(size=9)

        ws.row_dimensions[i].height = 18
        if status == "OPORTUNIDADE ALTA":
            opp_count += 1

    ws.freeze_panes = "A2"
    return opp_count

# ── ABA 1: DDD 34 ─────────────────────────────────────────────
ws1 = wb.active
n_opp_34 = escrever_aba(ws1, ddd34, "DDD 34 — Triângulo")

# ── ABA 2: TODOS MG ───────────────────────────────────────────
ws2 = wb.create_sheet("Todos MG")
n_opp_todos = escrever_aba(ws2, todos, "Todos MG")

# ── ABA 3: RESUMO ─────────────────────────────────────────────
ws3 = wb.create_sheet("Resumo")
ws3.title = "Resumo"

resumo_data = [
    ["MAPA DE OPORTUNIDADES — PCMSO TELEMEDICINA", "", ""],
    ["Higilabor + Dra. (tia) — Medicina do Trabalho telemedicina", "", ""],
    ["Data base: abril/2026", "", ""],
    ["", "", ""],
    ["STATUS", "QTD DDD 34", "DESCRIÇÃO"],
]

contagens_34 = {}
for r in ddd34:
    contagens_34[r["status"]] = contagens_34.get(r["status"], 0) + 1

for status in ["OPORTUNIDADE ALTA", "ITINERANTE", "COM CONCORRENTE", "SATURADO", "NÃO MAPEADO"]:
    qt = contagens_34.get(status, 0)
    desc = {
        "OPORTUNIDADE ALTA": "Sem provedor MT conhecido — entrar imediatamente",
        "ITINERANTE":        "Atendido por regional sem base local — competir com proximidade",
        "COM CONCORRENTE":   "Provedor local estabelecido — entrada mais difícil",
        "SATURADO":          "Múltiplas clínicas — não priorizar",
        "NÃO MAPEADO":       "Verificar CFM manualmente",
    }[status]
    resumo_data.append([status, qt, desc])

resumo_data += [
    ["", "", ""],
    ["COMO USAR ESTE MAPA", "", ""],
    ["1. OPORTUNIDADE ALTA → abordar imediatamente: apresentar PCMSO telemedicina + NR-1 psicossocial", "", ""],
    ["2. ITINERANTE → competir com tempo de resposta + relacionamento local", "", ""],
    ["3. NÃO MAPEADO → confirmar no CFM portal antes de abordar", "", ""],
    ["", "", ""],
    ["PREMISSA LEGAL (NR-7)", "", ""],
    ["Se não houver médico do trabalho registrado na localidade,", "", ""],
    ["outro especialista pode coordenar o PCMSO (via telemedicina conforme RES. CFM 2.314/2022)", "", ""],
    ["", "", ""],
    ["JANELA DE URGÊNCIA", "", ""],
    ["NR-1 psicossocial obrigatória a partir de 26/05/2026 — faltam 48 dias", "", ""],
    ["Abordagem: 'Seu PGR já contempla riscos psicossociais? A lei exige a partir de maio.'", "", ""],
]

for i, row in enumerate(resumo_data, 1):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=j, value=val)
        if i == 1:
            cell.font = Font(bold=True, size=14, color="1A237E")
        elif i == 5 and j <= 3:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A237E")
        elif i > 5 and i <= 5 + 5 and j == 1 and val in CORES:
            cell.fill = PatternFill("solid", fgColor=CORES[val])
            cell.font = Font(bold=True, color="FFFFFF")
        elif i > 5 and i <= 5 + 5 and j == 2:
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.font = Font(size=10)

ws3.column_dimensions["A"].width = 65
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 65

wb.save(OUTPUT)
print("\nArquivo salvo:", OUTPUT)
print("DDD 34 - OPORTUNIDADE ALTA:", n_opp_34, "municipios")
print("Todos MG - total mapeado:", len([r for r in todos if r['status'] != 'NAO MAPEADO']))

# Imprimir lista de oportunidades DDD 34
print("\n=== MUNICIPIOS DDD 34 - OPORTUNIDADE ALTA ===")
for r in ddd34:
    if r["status"] == "OPORTUNIDADE ALTA":
        tel = r['cel_prefeito'] or r['tel_municipio']
        nome = r['municipio'].encode('ascii', 'replace').decode()
        print(f"  {nome:30s} | {tel}")

print("\n=== MUNICIPIOS DDD 34 - ITINERANTE ===")
for r in ddd34:
    if r["status"] == "ITINERANTE":
        nome = r['municipio'].encode('ascii', 'replace').decode()
        obs = r['observacao'][:50].encode('ascii', 'replace').decode()
        print(f"  {nome:30s} | {obs}")
