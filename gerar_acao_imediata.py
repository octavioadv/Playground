"""
Gera planilha de ação imediata consolidada:
  Tab 1 — Empresas mineradoras do Pará (P.JURÍDICA) com email = prospecto SST
  Tab 2 — Municípios DDD 34 com oportunidade ALTA (sem MT local)
  Tab 3 — Resumo / script de abordagem
"""
import os, sys, re, unicodedata, openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def thin():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def norm(s):
    if not s: return ""
    s = str(s).strip().upper()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def cell_style(cell, bold=False, color="FFFFFF", bg=None, size=9, wrap=False, halign="left"):
    cell.font = Font(bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    cell.border = thin()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def header_row(ws, headers, widths, bg="1A237E"):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        cell_style(c, bold=True, color="FFFFFF", bg=bg, size=10, halign="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

# ─────────────────────────────────────────────────────────────
# 1. LER EMPRESAS DO PARÁ
# ─────────────────────────────────────────────────────────────
PARA_PATH = r'C:\Users\octav\OneDrive\Anexos\higilabor2.0'
para_file = None
for f in os.listdir(PARA_PATH):
    if 'Para' in f and f.endswith('.xlsx'):
        para_file = os.path.join(PARA_PATH, f)
        break

wb_in = openpyxl.load_workbook(para_file, read_only=True, data_only=True)
ws_pj = wb_in['P.JURÍDICA']

# Ler todas as linhas e juntar células fragmentadas (a planilha tem células quebradas)
raw_rows_pj = list(ws_pj.iter_rows(values_only=True))
wb_in.close()

def clean(v):
    if v is None: return ""
    return str(v).strip().replace("\n", " ").replace("\r", " ")

# Reconstruir registros: cada "empresa" ocupa várias linhas
# Estrutura: col0=vazio, col1=municipio, col2=empresa, col3=substância, col4=regime, col5=título/processo, col6=cnpj, col7=email
# Mas há linhas de continuação onde os valores estão fragmentados

# Estratégia: varrer linhas e agrupar por bloco (começa quando col6/CNPJ tem valor numérico ≥ 10 chars)
empresas = []
current = {"municipio": "", "empresa": "", "substancia": "", "regime": "", "titulo": "", "cnpj": "", "email": ""}

def flush(c):
    if c["empresa"].strip() and c["cnpj"].strip():
        empresas.append(dict(c))

def reset():
    return {"municipio": "", "empresa": "", "substancia": "", "regime": "", "titulo": "", "cnpj": "", "email": ""}

for i, row in enumerate(raw_rows_pj):
    r = [clean(v) for v in row] + [""] * 10
    # pular cabeçalho
    if i == 0 and ("MUNICÍPIO" in r[1].upper() or "TITULAR" in r[2].upper()):
        continue

    # detectar nova empresa: coluna CNPJ (col 6) tem valor que parece CNPJ
    cnpj_raw = re.sub(r'[.\-/]', '', r[6])
    if len(cnpj_raw) >= 11 and cnpj_raw.isdigit():
        flush(current)
        current = reset()
        current["cnpj"]       = cnpj_raw[:14]
        current["municipio"]  += (" " + r[1]).strip()
        current["empresa"]    += (" " + r[2]).strip()
        current["substancia"] += (" " + r[3]).strip()
        current["regime"]     += (" " + r[4]).strip()
        current["titulo"]     += (" " + r[5]).strip()
        current["email"]      += (" " + r[7]).strip()
    else:
        # linha de continuação: acumular campos não vazios
        if r[1]: current["municipio"]  += " " + r[1]
        if r[2]: current["empresa"]    += " " + r[2]
        if r[3]: current["substancia"] += " " + r[3]
        if r[4]: current["regime"]     += " " + r[4]
        if r[5]: current["titulo"]     += " " + r[5]
        if r[7]: current["email"]      += " " + r[7]

flush(current)  # último

# Limpar espaços duplos
for e in empresas:
    for k in e:
        e[k] = " ".join(e[k].split())

# ─────────────────────────────────────────────────────────────
# 2. CLASSIFICAR PRIORIDADE
# ─────────────────────────────────────────────────────────────
GRANDES = ["VALE", "ANGLO AMERICAN", "VOTORANTIM", "ALCOA", "NEXA", "XSTRATA",
           "CODELCO", "KINROSS", "MINERAÇÃO CARAÍBA", "ARAGUAIA NIQUEL"]

MINERIOS_CRITICOS = ["OURO", "COBRE", "NÍQUEL", "MANGANÊS"]  # NR-22 + silicose + PCMSO complexo
MINERIOS_SIMPLES  = ["CALCÁRIO", "AREIA", "BRITA", "ÁGUA MINERAL", "BAUXITA"]

def classificar(e):
    emp_up = e["empresa"].upper()
    sub_up = e["substancia"].upper()
    tem_email = bool(e["email"].strip())

    # grandes sem email útil
    for g in GRANDES:
        if g in emp_up:
            return "C - Baixa", "Grande empresa — provável SST próprio"

    # cooperativas e garimpeiros
    if any(x in emp_up for x in ["COOPERATIVA", "COOMIC", "GARIMPEIRO", "GARIMPO"]):
        return "C - Baixa", "Cooperativa/garimpeiro — ticket baixo"

    # sem email
    if not tem_email:
        return "C - Baixa", "Sem e-mail disponível"

    # mineração crítica com email
    for m in MINERIOS_CRITICOS:
        if m in sub_up:
            return "A - Alta", f"Mineral crítico ({m}) — NR-22 + PCMSO complexo + alto valor"

    for m in MINERIOS_SIMPLES:
        if m in sub_up:
            return "B - Média", f"Mineral simples ({m}) — PCMSO padrão"

    return "B - Média", "SST padrão mineração"

for e in empresas:
    e["prioridade"], e["justificativa"] = classificar(e)

# Separar email principal (primeiro endereço)
for e in empresas:
    emails = re.findall(r'[\w.\-+]+@[\w.\-]+', e["email"])
    e["email_principal"] = emails[0] if emails else ""
    e["emails_todos"] = " | ".join(emails)

# Ordenar: A primeiro, depois B, depois C
ordem = {"A - Alta": 0, "B - Média": 1, "C - Baixa": 2}
empresas.sort(key=lambda x: (ordem.get(x["prioridade"], 3), x["empresa"]))

# ─────────────────────────────────────────────────────────────
# 3. LER MUNICÍPIOS DDD 34 DE OPORTUNIDADE ALTA
# ─────────────────────────────────────────────────────────────
CLASSIFICACAO_DDD34 = {
    "ABADIA DOS DOURADOS":   "Sem provedor MT identificado",
    "ARAPORE":               "Sem provedor MT identificado",
    "CAMPO FLORIDO":         "Sem provedor MT identificado",
    "CARNEIRINHO":           "Sem provedor MT identificado",
    "CASCALHO RICO":         "Sem provedor MT identificado",
    "CENTRALINA":            "Ja cliente Higilabor — sem MT local",
    "COMENDADOR GOMES":      "Sem provedor MT identificado",
    "DELTA":                 "Sem provedor MT identificado",
    "DOURADOQUARA":          "Sem provedor MT identificado",
    "FRONTEIRA":             "Sem provedor MT identificado",
    "GRUPIARA":              "Sem provedor MT identificado",
    "GUIMARANIA":            "Sem provedor MT identificado",
    "IBIA":                  "Sem provedor MT identificado",
    "IPIACU":                "Sem provedor MT identificado",
    "IRAI DE MINAS":         "Sem provedor MT identificado",
    "LIMEIRA DO OESTE":      "Sem provedor MT identificado",
    "MATUTINA":              "Sem provedor MT identificado",
    "NOVA PONTE":            "Sem provedor MT identificado",
    "PEDRINOPOLIS":          "Sem provedor MT identificado",
    "PERDIZES":              "Sem provedor MT identificado",
    "PIRAJUBA":              "Sem provedor MT identificado",
    "PLANURA":               "Sem provedor MT identificado",
    "PRESIDENTE OLEGARIO":   "Sem provedor MT identificado",
    "RIO PARANAIBA":         "Sem provedor MT identificado",
    "ROMARIA":               "Sem provedor MT identificado",
    "SANTA JULIANA":         "Sem provedor MT identificado",
    "SAO FRANCISCO DE SALES":"Sem provedor MT identificado",
    "TAPIRA":                "Sem provedor MT identificado",
    "TUPACIGUARA":           "Sem provedor MT identificado",
    "UNIAO DE MINAS":        "Sem provedor MT identificado",
    "VARJAO DE MINAS":       "Sem provedor MT identificado",
    "VERISSIMO":             "Sem provedor MT identificado",
}

ONEDRIVE = r"C:\Users\octav\OneDrive\DOCS OCTAVIO"
municipios_file = None
for f in os.listdir(ONEDRIVE):
    if 'MUNIC' in f.upper() and f.endswith('.xlsx'):
        municipios_file = os.path.join(ONEDRIVE, f)
        break

wb_m = openpyxl.load_workbook(municipios_file, read_only=True, data_only=True)
ws_m = wb_m.active

municipios_opp = []
for i, row in enumerate(ws_m.iter_rows(values_only=True)):
    if i == 0: continue
    r = list(row) + [None] * 11
    num       = r[0]
    municipio = str(r[1]).strip() if r[1] else ""
    prefeito  = str(r[2]).strip() if r[2] else ""
    cel_pref  = str(r[3]).strip() if r[3] else ""
    tel_mun   = str(r[4]).strip() if r[4] else ""
    email     = str(r[5]).strip() if r[5] else ""
    if not municipio or municipio == "None": continue

    # detectar DDD
    ddd = ""
    for tel in [cel_pref, tel_mun]:
        s = str(tel).replace(" ","").replace("-","").replace("(","").replace(")","")
        if len(s) >= 2:
            ddd = s[:2]; break

    if ddd != "34": continue

    mun_norm = norm(municipio)
    if mun_norm in CLASSIFICACAO_DDD34:
        obs = CLASSIFICACAO_DDD34[mun_norm]
        municipios_opp.append({
            "municipio": municipio,
            "prefeito":  prefeito,
            "cel_pref":  cel_pref,
            "tel_mun":   tel_mun,
            "email_pref":email,
            "obs":       obs,
            "acao":      "Ligar para prefeito + apresentar PCMSO telemedicina + NR-1 psicossocial",
        })

wb_m.close()
municipios_opp.sort(key=lambda x: x["municipio"])

# ─────────────────────────────────────────────────────────────
# 4. GERAR EXCEL
# ─────────────────────────────────────────────────────────────
OUTPUT = r"C:\Users\octav\OneDrive\Apps\Claude\acao_imediata_higilabor.xlsx"
wb = Workbook()

# ─── ABA 1: PARA EMPRESAS ───────────────────────────────────
ws1 = wb.active
ws1.title = "Para - Mineradoras SST"

hdrs1 = ["Prioridade","Empresa","Município/PA","Substância","E-mail","CNPJ","Justificativa SST","Script de Abertura"]
wdts1 = [14, 38, 22, 22, 38, 18, 36, 55]
header_row(ws1, hdrs1, wdts1, bg="1B5E20")

PRIORIDADE_COR = {
    "A - Alta":  ("00C851", "FFFFFF"),
    "B - Média": ("FFBB33", "000000"),
    "C - Baixa": ("AAAAAA", "000000"),
}

scripts = {
    "A - Alta":
        "Bom dia! Sou da Higilabor SST. Vi que a [empresa] atua com mineração de [substância] no Pará. "
        "Com trabalhadores expostos a poeira e agentes químicos, o PCMSO e o PGR de mineração (NR-22) "
        "são obrigatórios. Além disso, a NR-1 psicossocial entra em vigor em maio/2026. "
        "Podemos fazer tudo por telemedicina + visita técnica. Consigo te mandar uma proposta hoje?",
    "B - Média":
        "Bom dia! Sou da Higilabor SST. A [empresa] precisa de PCMSO e PGR conforme NR-22 para sua "
        "operação de mineração. Trabalhamos com telemedicina — sem precisar de deslocamento constante. "
        "NR-1 psicossocial também entra em vigor em maio. Posso enviar proposta?",
    "C - Baixa":
        "Verificar se tem SST próprio antes de abordar.",
}

# Estatísticas
count_a = sum(1 for e in empresas if e["prioridade"] == "A - Alta")
count_b = sum(1 for e in empresas if e["prioridade"] == "B - Média")
count_c = sum(1 for e in empresas if e["prioridade"] == "C - Baixa")

for i, e in enumerate(empresas, 2):
    pri = e["prioridade"]
    bg_hex, fg_hex = PRIORIDADE_COR.get(pri, ("AAAAAA", "000000"))

    script = scripts.get(pri, "").replace("[empresa]", e["empresa"][:30]).replace("[substância]", e["substancia"][:20])

    vals = [
        pri, e["empresa"], e["municipio"][:40], e["substancia"][:30],
        e["email_principal"], e["cnpj"], e["justificativa"], script
    ]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.border = thin()
        c.font = Font(size=9, color="000000")
        c.alignment = Alignment(vertical="center", wrap_text=(col in [8]))
        if col == 1:
            c.fill = PatternFill("solid", fgColor=bg_hex)
            c.font = Font(bold=True, size=9, color=fg_hex)
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws1.row_dimensions[i].height = 16 if pri != "A - Alta" else 18

# ─── ABA 2: DDD 34 MUNICÍPIOS ───────────────────────────────
ws2 = wb.create_sheet("DDD 34 - Municípios sem MT")

hdrs2 = ["Município","Prefeito","Cel. Prefeito","Tel. Município","E-mail Prefeitura","Observação","Ação Imediata"]
wdts2 = [24, 24, 18, 18, 32, 32, 55]
header_row(ws2, hdrs2, wdts2, bg="0D47A1")

for i, m in enumerate(municipios_opp, 2):
    vals = [
        m["municipio"], m["prefeito"], m["cel_pref"],
        m["tel_mun"], m["email_pref"], m["obs"], m["acao"]
    ]
    bg_row = "E8F5E9" if m["municipio"] == "CENTRALINA" else "E3F2FD"
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.border = thin()
        c.font = Font(size=9)
        c.alignment = Alignment(vertical="center", wrap_text=(col == 7))
        c.fill = PatternFill("solid", fgColor=bg_row)
    ws2.row_dimensions[i].height = 16

# ─── ABA 3: RESUMO E SCRIPTS ────────────────────────────────
ws3 = wb.create_sheet("Resumo + Scripts")

resumo = [
    ["PLANO DE AÇÃO IMEDIATA — HIGILABOR SST", "", ""],
    ["Gerado em: abril/2026 | Responsável: Octav", "", ""],
    ["", "", ""],
    ["FRENTE 1: MINERADORAS DO PARÁ (via e-mail + telefone)", "", ""],
    ["Total de empresas mapeadas:", len(empresas), ""],
    ["  A - Alta prioridade (minerais críticos, email disponível):", count_a, "→ ABORDAR ESTA SEMANA"],
    ["  B - Média prioridade:", count_b, "→ Abordar na semana seguinte"],
    ["  C - Baixa prioridade:", count_c, "→ Não priorizar agora"],
    ["", "", ""],
    ["POR QUÊ MINERAÇÃO DO PARÁ?", "", ""],
    ["• NR-22 (Mineração): PCMSO + PGR + LTCAT obrigatórios com exigências específicas", "", ""],
    ["• Trabalhadores expostos a silica, metais pesados, ruído, calor extremo = PCMSO complexo + alto valor", "", ""],
    ["• NR-1 psicossocial: obrigatória a partir de 26/05/2026 — janela de urgência ABERTA", "", ""],
    ["• Muitas dessas empresas estão em cidades remotas sem SST local = oportunidade de telemedicina", "", ""],
    ["• Beto (parceiro Pará) pode prospectar presencialmente com respaldo técnico da Higilabor", "", ""],
    ["", "", ""],
    ["FRENTE 2: MUNICÍPIOS DDD 34 SEM MÉDICO DO TRABALHO LOCAL", "", ""],
    ["Total municípios de alta oportunidade:", len(municipios_opp), "→ Contato via prefeito/secretaria saúde"],
    ["• NR-7: onde não há MT na localidade, outro médico pode coordenar PCMSO", "", ""],
    ["• Dra. (tia) pode assinar ASOs e coordenar PCMSO por telemedicina (Res. CFM 2.314/2022)", "", ""],
    ["• Abertura: NR-1 psicossocial — 'A partir de maio, o PGR precisa contemplar psicossociais. Já adequaram?'", "", ""],
    ["", "", ""],
    ["SCRIPTS DE ABORDAGEM", "", ""],
    ["", "", ""],
    ["MINERADORAS (e-mail — Prioridade A):", "", ""],
    ["Assunto: PCMSO e PGR NR-22 para [empresa] — NR-1 Psicossocial obrigatória em maio/2026", "", ""],
    ["Corpo:", "", ""],
    ["Prezados,", "", ""],
    ["A Higilabor é especializada em SST para o setor de mineração. Identificamos que a [empresa]", "", ""],
    ["atua com mineração de [substância] e precisa de PCMSO, PGR (NR-22) e LTCAT atualizados.", "", ""],
    ["", "", ""],
    ["Novidade importante: a NR-1 psicossocial passa a ser obrigatória em 26/05/2026.", "", ""],
    ["Todas as empresas precisarão incluir avaliação de riscos psicossociais no PGR.", "", ""],
    ["", "", ""],
    ["Atuamos por telemedicina — sem necessidade de deslocamento constante.", "", ""],
    ["Podemos enviar uma proposta customizada esta semana?", "", ""],
    ["Att, Octav | Higilabor SST | (34) XXXX-XXXX", "", ""],
    ["", "", ""],
    ["MUNICÍPIOS DDD 34 (ligação para prefeito ou secretaria de saúde):", "", ""],
    ["'Bom dia, sou da Higilabor SST de Uberaba. Atendemos empresas em [município].", "", ""],
    ["A partir de maio, a NR-1 exige que o PGR inclua avaliação de riscos psicossociais.", "", ""],
    ["Trabalhamos com médica do trabalho por telemedicina — as empresas de vocês não precisam", "", ""],
    ["ir até Uberaba. Consigo montar uma lista das empresas que precisam regularizar?'", "", ""],
]

for i, row in enumerate(resumo, 1):
    for j, val in enumerate(row, 1):
        c = ws3.cell(row=i, column=j, value=val)
        if i == 1:
            c.font = Font(bold=True, size=14, color="1A237E")
        elif i in [4, 10, 17, 23]:
            c.font = Font(bold=True, size=11, color="0D47A1")
            c.fill = PatternFill("solid", fgColor="E3F2FD")
        elif str(val).startswith("  A - Alta") or str(val).startswith("→ ABORDAR"):
            c.font = Font(bold=True, size=10, color="00695C")
        elif j == 2 and isinstance(val, int):
            c.font = Font(bold=True, size=12)
            c.alignment = Alignment(horizontal="center")
        else:
            c.font = Font(size=9)

ws3.column_dimensions["A"].width = 70
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 30

wb.save(OUTPUT)

print("Salvo:", OUTPUT)
print(f"\nRESUMO:")
print(f"  Mineradoras Para - A (alta): {count_a}")
print(f"  Mineradoras Para - B (media): {count_b}")
print(f"  Mineradoras Para - C (baixa): {count_c}")
print(f"  Municipios DDD 34 oportunidade alta: {len(municipios_opp)}")
print(f"\nTop 10 mineradoras Prioridade A:")
for e in [x for x in empresas if x['prioridade'] == 'A - Alta'][:10]:
    nome = e['empresa'][:40].encode('ascii','replace').decode()
    email = e['email_principal']
    print(f"  {nome:42s}  {email}")
